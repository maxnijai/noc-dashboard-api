"""Online SA1-4 SLA Improvement - retrospective analysis of CLOSED Online
SA1-4 tickets, imported from a manually-uploaded Excel export (not the
live Google Sheet everything else in this app reads from - this dataset
is historical/closed, refreshed by a daily manual import, held in memory
between imports like every other in-memory cache here).

Data source note: TICKET_SLA ("within"/"over") is a PRE-COMPUTED outcome
field from the source system - this module reuses it exactly as given,
it does not recompute an SLA verdict from TARGETFINISH (that's a
different, live-ticket concept used by the P0/P1/P2 priority formula
elsewhere in this app, which doesn't apply to already-closed tickets).
"""

import logging
import os
import pickle
import re
import threading
from datetime import datetime

import openpyxl

log = logging.getLogger(__name__)

REQUIRED_COLUMNS = [
    "TICKETID", "CREATIONDATE", "TARGETFINISH", "TICKET_SLA",
    "TRUEOWNERGROUP", "PROVINCE_EN", "PROBLEM", "SUB_CAUSE", "CI_Name",
]

# TRUEOWNERGROUP -> Region + Province, e.g. "TRUE-TH-BBT-NOR1-CMI1-NOP"
# -> region "NOR1", code "CMI1". Validated against the actual uploaded
# file: 100% of 12,697 rows matched this pattern, so it's used directly
# rather than falling back to PROVINCE_EN (which the user's spec asked to
# NOT use for this - PROVINCE_EN would silently merge CMI1/CMI2 without
# the explicit, visible merge step the spec asked for).
_TOG_PATTERN = re.compile(r"TRUE-TH-BBT-(NOR[12])-([A-Z0-9]+)-NOP")
# Only known 2-zone province in the data - every other code maps to itself.
_PROVINCE_CODE_MERGE = {"CMI1": "CMI", "CMI2": "CMI"}


def _extract_region_province(true_owner_group):
    m = _TOG_PATTERN.match(str(true_owner_group or "").strip())
    if not m:
        return None, None
    region, code = m.group(1), m.group(2)
    province = _PROVINCE_CODE_MERGE.get(code, code)
    return region, province


def _parse_excel_dt(v):
    """Cell values come back as datetime objects already when the sheet
    stores them as real dates (openpyxl with data_only=True) - falls back
    to string parsing only if the cell was stored as text."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _iso_week_label(dt):
    iso_year, iso_week, _ = dt.isocalendar()
    return f"{iso_year}-W{iso_week:02d}"


def _normalize_text(s):
    """Collapses irregular whitespace (non-breaking spaces etc - confirmed
    present in this file's SUB_CAUSE values) before using a field as a
    grouping key, so two values that look identical don't end up in
    separate groups."""
    return re.sub(r"\s+", " ", str(s or "")).strip()


class ImportValidationError(Exception):
    """Raised when the uploaded file is missing required columns or has
    no usable rows - carries a user-facing message, never lets a bad
    import silently replace good data (see import_excel_file)."""
    pass


def parse_excel_file(file_path):
    """Reads the uploaded .xlsx, validates required columns are present,
    and returns (rows, warnings) - rows are enriched with parsed dates,
    Region/Province, and ISO week/day labels. Never raises for row-level
    data issues (a bad row is skipped and counted in warnings); raises
    ImportValidationError only for structural problems (missing columns,
    zero usable rows) that would make the whole import meaningless."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        raise ImportValidationError(
            f"เปิดไฟล์ไม่ได้ ({e}) - กรุณาตรวจสอบว่าเป็นไฟล์ .xlsx ที่ไม่เสียหาย แล้วลองใหม่"
        )
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ImportValidationError("ไฟล์ว่างเปล่า ไม่มีข้อมูล")
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    header_idx = {h: i for i, h in enumerate(headers)}

    missing = [c for c in REQUIRED_COLUMNS if c not in header_idx]
    if missing:
        raise ImportValidationError(
            f"ไฟล์ขาด Column ที่จำเป็น: {', '.join(missing)} — กรุณาตรวจสอบไฟล์แล้วลองใหม่"
        )

    warnings = []
    skipped_missing_fields = 0
    skipped_bad_dates = 0
    skipped_bad_region = 0
    skipped_bad_sla = 0
    seen_ticket_ids = set()
    duplicate_ticket_ids = 0

    rows = []
    for raw in rows_iter:
        get = lambda col: raw[header_idx[col]] if col in header_idx and header_idx[col] < len(raw) else None

        ticket_id = str(get("TICKETID") or "").strip()
        if not ticket_id:
            skipped_missing_fields += 1
            continue
        if ticket_id in seen_ticket_ids:
            duplicate_ticket_ids += 1
            continue
        seen_ticket_ids.add(ticket_id)

        creation_dt = _parse_excel_dt(get("CREATIONDATE"))
        target_dt = _parse_excel_dt(get("TARGETFINISH"))
        if not creation_dt or not target_dt:
            skipped_bad_dates += 1
            continue

        ticket_sla = str(get("TICKET_SLA") or "").strip().lower()
        if ticket_sla not in ("within", "over"):
            skipped_bad_sla += 1
            continue

        region, province = _extract_region_province(get("TRUEOWNERGROUP"))
        if not region or not province:
            skipped_bad_region += 1
            continue

        rows.append({
            "TICKETID": ticket_id,
            "CREATIONDATE": creation_dt,
            "TARGETFINISH": target_dt,
            "CLOSEDTIME": _parse_excel_dt(get("CLOSEDTIME")),
            "TICKET_SLA": ticket_sla,  # 'within' | 'over'
            "TRUEOWNERGROUP": str(get("TRUEOWNERGROUP") or "").strip(),
            "region": region,
            "province": province,
            "province_en_raw": str(get("PROVINCE_EN") or "").strip(),
            "PROBLEM": _normalize_text(get("PROBLEM")) or "(ไม่ระบุ)",
            "SUB_CAUSE": _normalize_text(get("SUB_CAUSE")) or "(ไม่ระบุ)",
            "CI_Name": str(get("CI_Name") or "").strip() or "(ไม่ระบุ)",
            "DISTRICT_EN": str(get("DISTRICT_EN") or "").strip(),  # optional column - blank/missing handled at analysis time (see _normalize_district), never required for import to succeed
            "iso_week": _iso_week_label(creation_dt),
            "iso_date": creation_dt.date().isoformat(),
        })

    if not rows:
        raise ImportValidationError(
            "ไม่พบข้อมูลที่ใช้งานได้เลยในไฟล์นี้ (อาจเป็นเพราะ Format วันที่ผิด หรือ TICKET_SLA ไม่ใช่ within/over) — Dashboard เดิมจะไม่ถูกแทนที่"
        )

    if skipped_missing_fields:
        warnings.append(f"ข้าม {skipped_missing_fields} แถวที่ไม่มี TICKETID")
    if duplicate_ticket_ids:
        warnings.append(f"พบ TICKETID ซ้ำ {duplicate_ticket_ids} รายการ - ใช้รายการแรกที่เจอ")
    if skipped_bad_dates:
        warnings.append(f"ข้าม {skipped_bad_dates} แถวที่ CREATIONDATE หรือ TARGETFINISH อ่านไม่ได้")
    if skipped_bad_sla:
        warnings.append(f"ข้าม {skipped_bad_sla} แถวที่ TICKET_SLA ไม่ใช่ within/over")
    if skipped_bad_region:
        warnings.append(f"ข้าม {skipped_bad_region} แถวที่ TRUEOWNERGROUP ไม่ตรงรูปแบบที่รู้จัก (TRUE-TH-BBT-NOR1/2-XXX-NOP)")

    return rows, warnings


# ── In-memory store, backed by a disk snapshot for durability across
# process restarts (deploys, crashes, Railway idle-sleep) - the store
# itself is still what every read goes through (module-level dict, exactly
# like every other cache in this app), the disk file is only ever
# written to on a successful import and read once at startup, so normal
# operation never touches disk. Data is only ever REPLACED wholesale by
# the next successful import, never partially updated. ───────────────
_store = {"rows": None, "warnings": [], "imported_at": None, "filename": None}
_store_lock = threading.Lock()


def _resolve_store_path():
    """Prefers a Railway persistent volume mounted at /data (the
    conventional path Railway's docs suggest) if one is actually writable;
    falls back to a file next to this module (survives process restarts
    within the same container, though not a fresh redeploy unless a
    volume is attached) - returns None only if neither location is
    writable, in which case persistence is silently disabled and the
    dataset behaves exactly as it did before (in-memory only)."""
    candidates = ["/data/sla_improvement_store.pkl", os.path.join(os.path.dirname(os.path.abspath(__file__)), "sla_improvement_store.pkl")]
    for path in candidates:
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            probe = path + ".probe"
            with open(probe, "w") as f:
                f.write("probe")
            os.remove(probe)
            return path
        except OSError:
            continue
    return None


_STORE_FILE_PATH = _resolve_store_path()


def _save_store_to_disk():
    if not _STORE_FILE_PATH:
        return
    try:
        with _store_lock:
            snapshot = dict(_store)
        tmp_path = _STORE_FILE_PATH + ".tmp"
        with open(tmp_path, "wb") as f:
            pickle.dump(snapshot, f)
        os.replace(tmp_path, _STORE_FILE_PATH)  # atomic swap - a crash mid-write never corrupts the last good snapshot
    except Exception:
        log.exception("Failed to persist SLA Improvement data to disk - it will still work for the rest of this process's lifetime, just won't survive a restart")


def _load_store_from_disk():
    if not _STORE_FILE_PATH or not os.path.exists(_STORE_FILE_PATH):
        return
    try:
        with open(_STORE_FILE_PATH, "rb") as f:
            loaded = pickle.load(f)
        with _store_lock:
            _store.update(loaded)
        log.info(f"Loaded persisted SLA Improvement data from disk: {len(loaded.get('rows') or [])} rows, imported_at={loaded.get('imported_at')}")
    except Exception:
        log.exception("Failed to load persisted SLA Improvement data from disk - starting empty")


_load_store_from_disk()  # attempt once at module import (app startup)


def import_excel_file(file_path, filename):
    """Parses + validates the file, and ONLY replaces the store (memory
    AND the disk snapshot) if parsing succeeds end to end - a bad import
    raises ImportValidationError and leaves whatever was there before
    untouched, so "Dashboard เดิมจะไม่ถูกแทนที่" (the existing dashboard
    is never broken by a bad upload) holds even at this layer, not just
    in the error message."""
    rows, warnings = parse_excel_file(file_path)  # raises ImportValidationError on structural problems
    with _store_lock:
        _store["rows"] = rows
        _store["warnings"] = warnings
        _store["imported_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        _store["filename"] = filename
    _save_store_to_disk()
    return {"row_count": len(rows), "warnings": warnings}


def get_store_status():
    with _store_lock:
        return {
            "has_data": _store["rows"] is not None,
            "row_count": len(_store["rows"]) if _store["rows"] else 0,
            "imported_at": _store["imported_at"],
            "filename": _store["filename"],
            "warnings": _store["warnings"],
        }


def get_rows():
    """Public accessor for the currently-imported dataset (or None if
    nothing's been imported yet) - callers outside this module should use
    this instead of reaching into _store directly."""
    with _store_lock:
        return _store["rows"]


def _pct(numerator, denominator):
    return round(numerator / denominator * 100, 2) if denominator else None


def _split_over(rows):
    return len(rows), sum(1 for r in rows if r["TICKET_SLA"] == "over")


# ── Province ranking (🔴 Worst / 🟠 At Risk / 🔵 Good) ──────────────────

def _filter_last_n_days(rows, n):
    """Rows whose iso_date falls among the most recent N distinct dates
    present in the dataset - "recent" is anchored to the latest date
    actually in the data (same convention as War Room's "today"), not the
    literal calendar date, since this is a manually-imported historical
    dataset that may lag behind or be backfilled."""
    all_dates = sorted({r["iso_date"] for r in rows})
    recent_dates = set(all_dates[-n:])
    return [r for r in rows if r["iso_date"] in recent_dates]


def build_province_ranking(rows):
    total_all, over_all = _split_over(rows)
    overall_pct = _pct(over_all, total_all) or 0

    by_province = {}
    for r in rows:
        by_province.setdefault(r["province"], []).append(r)

    ranking = []
    for prov, prov_rows in by_province.items():
        total, over = _split_over(prov_rows)
        pct_over = _pct(over, total)
        region = prov_rows[0]["region"]
        if pct_over is None:
            tier = None
        elif pct_over > overall_pct * 1.25:
            tier = "worst"
        elif pct_over > overall_pct:
            tier = "at_risk"
        else:
            tier = "good"
        ranking.append({
            "province": prov, "region": region, "total": total, "over": over,
            "pct_over": pct_over, "tier": tier,
        })
    ranking.sort(key=lambda r: -(r["pct_over"] or 0))
    return {"overall_pct_over": overall_pct, "rows": ranking}


# ── Trend (Daily / Weekly x Overall / Region / Province) ───────────────

def _trend_series(rows, date_key):
    buckets = {}
    for r in rows:
        buckets.setdefault(r[date_key], []).append(r)
    series = []
    for period in sorted(buckets.keys()):
        total, over = _split_over(buckets[period])
        within = total - over
        series.append({
            "period": period, "total": total, "over": over, "within": within,
            "pct_over": _pct(over, total), "pct_within": _pct(within, total),
        })
    return series


def _with_rolling_avg(series, window=7):
    """Adds a trailing N-period rolling average of pct_over to each point
    (e.g. the "Avg 7D% over" line) - trailing only (never looks ahead),
    and only computed once enough prior periods exist to average over,
    matching how a 7-day rolling average is normally read on a chart."""
    out = []
    for i, p in enumerate(series):
        window_pts = series[max(0, i - window + 1):i + 1]
        vals = [w["pct_over"] for w in window_pts if w["pct_over"] is not None]
        rolling = round(sum(vals) / len(vals), 2) if vals else None
        out.append({**p, "rolling_avg_pct_over": rolling})
    return out


def build_trend(rows):
    """Returns daily and weekly series, each with overall + by-region +
    by-province breakdowns. Precomputed for every province/region since
    this only runs once per import, not per page view. The "overall"
    series also carries a 7-period rolling average of %Over (daily = 7
    day rolling, weekly = 7 week rolling), matching the reference chart's
    "Avg 7D% over" line."""
    result = {"daily": {"overall": _with_rolling_avg(_trend_series(rows, "iso_date")), "by_region": {}, "by_province": {}},
              "weekly": {"overall": _with_rolling_avg(_trend_series(rows, "iso_week")), "by_region": {}, "by_province": {}}}
    regions = sorted({r["region"] for r in rows})
    provinces = sorted({r["province"] for r in rows})
    for reg in regions:
        reg_rows = [r for r in rows if r["region"] == reg]
        result["daily"]["by_region"][reg] = _trend_series(reg_rows, "iso_date")
        result["weekly"]["by_region"][reg] = _trend_series(reg_rows, "iso_week")
    for prov in provinces:
        prov_rows = [r for r in rows if r["province"] == prov]
        result["daily"]["by_province"][prov] = _trend_series(prov_rows, "iso_date")
        result["weekly"]["by_province"][prov] = _trend_series(prov_rows, "iso_week")
    return result


def build_improvement_heatmap(rows, daily_periods=14):
    """Two heatmaps (daily + weekly) of %Over SLA by Province, meant to
    visualize improvement/regression over time at a glance - daily is
    capped to the most recent `daily_periods` days (per explicit request,
    to keep it readable) so it doesn't sprawl across months; weekly shows
    every ISO week present in the data (naturally a much shorter list)."""
    provinces = sorted({r["province"] for r in rows}, key=lambda p: -sum(1 for r in rows if r["province"] == p))
    all_daily_periods = sorted({r["iso_date"] for r in rows})
    recent_daily_periods = all_daily_periods[-daily_periods:]
    all_weekly_periods = sorted({r["iso_week"] for r in rows})

    def _matrix(ordered_periods, date_key):
        period_set = set(ordered_periods)  # fast membership check only - iteration must use the ORDERED list below, or cell order won't match the periods header
        by_province_period = {}
        for r in rows:
            period = r[date_key]
            if period not in period_set:
                continue
            key = (r["province"], period)
            by_province_period.setdefault(key, {"total": 0, "over": 0})
            by_province_period[key]["total"] += 1
            if r["TICKET_SLA"] == "over":
                by_province_period[key]["over"] += 1
        rows_out = []
        for prov in provinces:
            cells = []
            for period in ordered_periods:
                cell = by_province_period.get((prov, period))
                cells.append({"period": period, "total": cell["total"] if cell else 0, "over": cell["over"] if cell else 0, "pct_over": _pct(cell["over"], cell["total"]) if cell else None})
            rows_out.append({"province": prov, "region": next(r["region"] for r in rows if r["province"] == prov), "cells": cells})
        return rows_out

    return {
        "daily": {"periods": recent_daily_periods, "rows": _matrix(recent_daily_periods, "iso_date")},
        "weekly": {"periods": all_weekly_periods, "rows": _matrix(all_weekly_periods, "iso_week")},
    }


# ── Root cause drill-down (Problem -> Sub-Cause -> CI -> Province) ─────

def _ranked_group(scoped_rows, key, top_n):
    """scoped_rows: rows already filtered to the current drill-down scope
    (by Problem/Sub-Cause/etc as needed), but NOT pre-filtered to
    TICKET_SLA=='over' - both within and over tickets must be included
    here, or pct_over per group is meaningless (it would trivially be
    100% for every group, since "over count / over count" is always 1).
    Ranks by OVER count (what actually matters for "what's causing Over
    SLA"), while total/pct_over reflect each group's own full ticket
    volume within the current scope."""
    total_over_in_scope = sum(1 for r in scoped_rows if r["TICKET_SLA"] == "over")
    counts = {}
    for r in scoped_rows:
        counts.setdefault(r[key], {"total": 0, "over": 0})
        counts[r[key]]["total"] += 1
        if r["TICKET_SLA"] == "over":
            counts[r[key]]["over"] += 1
    ranked = [
        {"label": k, "total": v["total"], "over": v["over"], "pct_over": _pct(v["over"], v["total"]),
         "share_of_total_over": _pct(v["over"], total_over_in_scope)}
        for k, v in counts.items() if v["over"] > 0
    ]
    ranked.sort(key=lambda r: -r["over"])
    return ranked[:top_n]


def build_root_cause(rows, top_n=15):
    total_over_all = sum(1 for r in rows if r["TICKET_SLA"] == "over")
    return {
        "total_over": total_over_all,
        "top_problems": _ranked_group(rows, "PROBLEM", top_n),
        "top_sub_causes": _ranked_group(rows, "SUB_CAUSE", top_n),
        "top_ci": _ranked_group(rows, "CI_Name", top_n),
    }


def drill_down_root_cause(rows, problem=None, sub_cause=None, ci_name=None, top_n=15):
    """One level of drill-down at a time - Problem -> Sub-Cause -> CI ->
    Province -> Ticket. Called with whichever filters are already
    selected; returns the next level's ranked breakdown plus the
    matching ticket list at the deepest level. Filters are applied to the
    FULL row set (within + over), not just over-SLA rows, so each
    level's own pct_over stays meaningful - see _ranked_group."""
    scoped_rows = rows
    if problem:
        scoped_rows = [r for r in scoped_rows if r["PROBLEM"] == problem]
    if sub_cause:
        scoped_rows = [r for r in scoped_rows if r["SUB_CAUSE"] == sub_cause]
    if ci_name:
        scoped_rows = [r for r in scoped_rows if r["CI_Name"] == ci_name]
    total_over_in_scope = sum(1 for r in scoped_rows if r["TICKET_SLA"] == "over")

    result = {"matched_over_tickets": total_over_in_scope}
    if not problem:
        result["next_level"] = "PROBLEM"
        result["options"] = _ranked_group(scoped_rows, "PROBLEM", top_n)
    elif not sub_cause:
        result["next_level"] = "SUB_CAUSE"
        result["options"] = _ranked_group(scoped_rows, "SUB_CAUSE", top_n)
    elif not ci_name:
        result["next_level"] = "CI_Name"
        result["options"] = _ranked_group(scoped_rows, "CI_Name", top_n)
    else:
        result["next_level"] = "province"
        result["options"] = _ranked_group(scoped_rows, "province", top_n)
        over_scoped = [r for r in scoped_rows if r["TICKET_SLA"] == "over"]
        result["tickets"] = [
            {"TICKETID": r["TICKETID"], "province": r["province"], "region": r["region"],
             "CREATIONDATE": r["CREATIONDATE"].strftime("%Y-%m-%d %H:%M:%S"),
             "TARGETFINISH": r["TARGETFINISH"].strftime("%Y-%m-%d %H:%M:%S")}
            for r in over_scoped[:200]  # cap - this is a drill-down detail list, not a full export
        ]
    return result


# ── Impact / Risk classification (never % alone - see module docstring
# in the request this implements: count AND % together) ────────────────

def _classify_impact_risk(entities, total_over_all, overall_pct_over):
    """entities: list of dicts with 'total'/'over'/'pct_over' already set.
    Adds 'impact_share' (this entity's % of ALL over-SLA tickets) and
    'classification' (priority_improvement / high_impact / high_risk /
    normal) - thresholds are DATA-DRIVEN (median impact share, and the
    dataset's own overall %Over SLA as the risk bar), not fixed constants,
    so they adapt to whatever's actually in the imported data."""
    for e in entities:
        e["impact_share"] = _pct(e["over"], total_over_all) or 0
    impact_shares = sorted(e["impact_share"] for e in entities if e["over"] > 0)
    median_impact = impact_shares[len(impact_shares) // 2] if impact_shares else 0
    for e in entities:
        is_high_impact = e["impact_share"] >= median_impact and e["over"] > 0
        is_high_risk = (e["pct_over"] or 0) >= overall_pct_over
        if is_high_impact and is_high_risk:
            e["classification"] = "priority_improvement"
        elif is_high_impact:
            e["classification"] = "high_impact"
        elif is_high_risk:
            e["classification"] = "high_risk"
        else:
            e["classification"] = "normal"
    return entities


def build_impact_risk(rows):
    total_all, over_all = _split_over(rows)
    overall_pct = _pct(over_all, total_all) or 0
    by_province = {}
    for r in rows:
        by_province.setdefault(r["province"], []).append(r)
    entities = []
    for prov, prov_rows in by_province.items():
        total, over = _split_over(prov_rows)
        entities.append({"province": prov, "region": prov_rows[0]["region"], "total": total, "over": over, "pct_over": _pct(over, total)})
    entities = _classify_impact_risk(entities, over_all, overall_pct)
    entities.sort(key=lambda e: -e["impact_share"])
    return entities


# ── Executive KPI + War Room + Management table ─────────────────────────

def build_executive_kpi(rows, province_ranking, root_cause):
    total, over = _split_over(rows)
    pct_over = _pct(over, total)

    by_region = {}
    for r in rows:
        by_region.setdefault(r["region"], []).append(r)
    region_stats = []
    for reg, reg_rows in by_region.items():
        t, o = _split_over(reg_rows)
        region_stats.append({"region": reg, "total": t, "over": o, "pct_over": _pct(o, t)})
    region_stats.sort(key=lambda r: r["pct_over"] or 0)
    best_region = region_stats[0] if region_stats else None
    worst_region = region_stats[-1] if region_stats else None

    prov_rows = province_ranking["rows"]
    best_province = min(prov_rows, key=lambda r: r["pct_over"] if r["pct_over"] is not None else 999) if prov_rows else None
    worst_province = max(prov_rows, key=lambda r: r["pct_over"] if r["pct_over"] is not None else -1) if prov_rows else None

    top_opportunity = None
    if root_cause["top_problems"]:
        p = root_cause["top_problems"][0]
        top_opportunity = f"{p['label']} ({p['over']} ticket, {p['share_of_total_over']}% ของ Over SLA ทั้งหมด)"

    return {
        "total": total, "over": over, "pct_over": pct_over,
        "best_region": best_region, "worst_region": worst_region,
        "best_province": best_province, "worst_province": worst_province,
        "top_improvement_opportunity": top_opportunity,
    }


def build_war_room(rows, root_cause, province_ranking):
    """"Today" = the most recent CREATIONDATE day present in the imported
    data, not the literal calendar date - this dataset is a manual daily
    import that may lag behind or include backfilled history, so "latest
    day actually in the data" is the meaningful anchor, clearly labeled
    as such rather than assumed to be the literal today."""
    if not rows:
        return None
    latest_date = max(r["iso_date"] for r in rows)
    today_rows = [r for r in rows if r["iso_date"] == latest_date]
    total, over = _split_over(today_rows)

    today_root_cause = build_root_cause(today_rows, top_n=1)
    top_problem = today_root_cause["top_problems"][0] if today_root_cause["top_problems"] else None
    top_province_today = None
    if today_rows:
        by_prov = {}
        for r in today_rows:
            by_prov.setdefault(r["province"], []).append(r)
        prov_over = [(p, sum(1 for x in rs if x["TICKET_SLA"] == "over")) for p, rs in by_prov.items()]
        prov_over = [x for x in prov_over if x[1] > 0]
        if prov_over:
            prov_over.sort(key=lambda x: -x[1])
            top_province_today = {"province": prov_over[0][0], "over": prov_over[0][1]}
    top_ci_today = today_root_cause["top_ci"][0] if today_root_cause["top_ci"] else None

    # Improvement focus: the single highest-impact Priority Improvement
    # province overall (not just today - a one-day snapshot is too noisy
    # to drive a "what to fix" call), so War Room can point at something
    # backed by the full imported window.
    priority_provinces = [p for p in province_ranking["rows"] if p["tier"] == "worst"]
    improvement_focus = None
    if priority_provinces:
        p = priority_provinces[0]
        improvement_focus = f"{p['province']} ({p['region']}) - {p['pct_over']}% Over SLA จาก {p['total']} ticket"

    return {
        "as_of_date": latest_date,
        "today_total": total, "today_over": over, "today_pct_over": _pct(over, total),
        "top_problem_today": top_problem,
        "top_province_today": top_province_today,
        "top_ci_today": top_ci_today,
        "improvement_focus": improvement_focus,
    }


def build_management_table(province_ranking, impact_risk):
    """Region/Province/Total/Over/%Over/Rank/Impact - merges the ranking
    (for %Over/tier) with the impact/risk classification (for the Impact
    label), sorted by %Over SLA descending by default; the frontend
    handles re-sorting by any column."""
    impact_by_province = {e["province"]: e["classification"] for e in impact_risk}
    rows = []
    for i, r in enumerate(province_ranking["rows"]):
        rows.append({
            "region": r["region"], "province": r["province"],
            "total": r["total"], "over": r["over"], "pct_over": r["pct_over"],
            "rank": i + 1, "impact": impact_by_province.get(r["province"], "normal"),
        })
    return rows


# ── Province Deep Dive (District-level analysis) ────────────────────────
# New analysis layer, additive only - does not modify any function above.
# Confirmed against the real imported data: DISTRICT_EN has 172 distinct
# values, clean formatting (no case/whitespace-variant duplicates), only
# ~0.04% of rows have it missing (shown as "Unknown / N/A" here, counted,
# never dropped). Five district names are reused across DIFFERENT
# provinces (a real Thailand naming collision, e.g. "SARAPHI" exists in
# both CMI and LPN) - every function below scopes to ONE province FIRST,
# then groups by district WITHIN that already-filtered subset, so this
# collision can never merge two unrelated districts together.

UNKNOWN_DISTRICT = "Unknown / N/A"


def _normalize_district(raw):
    """Display-time only - never touches source data. A blank/null cell
    becomes "Unknown / N/A" (explicit, counted) rather than being dropped
    or silently merged into some other bucket."""
    s = _normalize_text(raw)
    return s if s else UNKNOWN_DISTRICT


def _classify_district_tier(impact_share, pct_over, overall_pct_over, median_impact_share):
    """🔴 Critical/High Impact / 🟠 High Risk / 🟡 Monitor / 🔵 Good - same
    two-dimensional idea as _classify_impact_risk (never % alone), just
    the label set this specific request asked for. Thresholds are
    data-driven: overall %Over SLA as the risk bar, median impact share
    among districts that have at least one Over SLA ticket - not fixed
    constants, so they adapt to whatever the current import contains."""
    is_high_impact = impact_share >= median_impact_share and impact_share > 0
    is_high_risk = (pct_over or 0) >= overall_pct_over
    if is_high_impact and is_high_risk:
        return "critical"
    if is_high_impact:
        return "high_impact"
    if is_high_risk:
        return "high_risk"
    return "monitor" if (pct_over or 0) > 0 else "good"


def build_district_table(province_rows, overall_pct_over):
    """District breakdown WITHIN an already-province-scoped row set -
    Total/Over/%Over/Impact tier, sorted Impact-first (per explicit
    request), the frontend re-sorts by any column from there."""
    total_all, over_all = _split_over(province_rows)
    by_district = {}
    for r in province_rows:
        d = _normalize_district(r["DISTRICT_EN"])
        by_district.setdefault(d, []).append(r)

    entities = []
    for d, d_rows in by_district.items():
        total, over = _split_over(d_rows)
        entities.append({"district": d, "total": total, "over": over, "pct_over": _pct(over, total)})

    impact_shares = sorted((_pct(e["over"], over_all) or 0) for e in entities if e["over"] > 0)
    median_impact = impact_shares[len(impact_shares) // 2] if impact_shares else 0
    for e in entities:
        e["impact_share"] = _pct(e["over"], over_all) or 0
        e["tier"] = _classify_district_tier(e["impact_share"], e["pct_over"], overall_pct_over, median_impact)

    # Impact-first default order: by over-count (the concrete magnitude),
    # not by %, matching "อย่าใช้ % อย่างเดียว" (never rank by % alone).
    entities.sort(key=lambda e: -e["over"])
    for i, e in enumerate(entities):
        e["rank"] = i + 1
    return entities


def _pareto(entities, label_key, value_key="over"):
    """Top-N + cumulative share, generic across whatever level is passed
    in (District/Problem/Sub-Cause/CI) - all computed from real data, no
    hardcoded category list at any level."""
    total = sum(e[value_key] for e in entities)
    ordered = sorted(entities, key=lambda e: -e[value_key])
    cumulative = 0
    out = []
    for e in ordered:
        cumulative += e[value_key]
        out.append({
            "label": e[label_key], value_key: e[value_key],
            "share": _pct(e[value_key], total), "cumulative_share": _pct(cumulative, total),
        })
    return out


def build_province_deep_dive(rows, province_ranking, province):
    """Everything for one province: performance vs region average, trend,
    district table (+ pareto), root-cause top lists scoped to this
    province, an auto-generated insight, and a short Improvement
    Opportunity list. rows is the FULL dataset - this function does its
    own province filtering, so callers never need to pre-filter."""
    province_rows = [r for r in rows if r["province"] == province]
    if not province_rows:
        return None
    region = province_rows[0]["region"]
    total, over = _split_over(province_rows)
    pct_over = _pct(over, total)

    region_rows = [r for r in rows if r["region"] == region]
    region_total, region_over = _split_over(region_rows)
    region_pct_over = _pct(region_over, region_total)

    prov_rank_row = next((r for r in province_ranking["rows"] if r["province"] == province), None)
    rank = province_ranking["rows"].index(prov_rank_row) + 1 if prov_rank_row else None

    daily_trend = _trend_series(province_rows, "iso_date")
    weekly_trend = _trend_series(province_rows, "iso_week")

    district_table = build_district_table(province_rows, pct_over or 0)
    district_pareto = _pareto(district_table, "district")

    over_province_rows = [r for r in province_rows if r["TICKET_SLA"] == "over"]
    top_problems = _ranked_group(province_rows, "PROBLEM", 10)
    top_sub_causes = _ranked_group(province_rows, "SUB_CAUSE", 10)
    top_ci = _ranked_group(province_rows, "CI_Name", 10)
    problem_pareto = _pareto([{"label": p["label"], "over": p["over"]} for p in top_problems], "label")
    sub_cause_pareto = _pareto([{"label": p["label"], "over": p["over"]} for p in top_sub_causes], "label")
    ci_pareto = _pareto([{"label": p["label"], "over": p["over"]} for p in top_ci], "label")

    top_district = district_table[0] if district_table else None
    top_problem = top_problems[0] if top_problems else None
    top_sub_cause = top_sub_causes[0] if top_sub_causes else None
    top_ci_item = top_ci[0] if top_ci else None

    # Auto-generated insight text - every number below comes directly
    # from the values just computed, nothing hardcoded.
    is_drag = (pct_over or 0) > (region_pct_over or 0)
    key_finding = (
        f"{province} มี Over SLA {over} Ticket จากทั้งหมด {total} Ticket คิดเป็น {pct_over}% "
        f"ซึ่ง{'สูงกว่า' if is_drag else 'ต่ำกว่า'}ค่าเฉลี่ย {region} ที่ {region_pct_over}%"
    )
    focus_point = None
    if top_district and top_district["over"] > 0:
        focus_point = f"เขต {top_district['district']} เป็นจุดที่มี Impact สูงสุด คิดเป็น {top_district['impact_share']}% ของ Over SLA ใน {province}"
    root_cause_text = None
    if top_problem:
        root_cause_text = f"ปัญหาหลักคือ {top_problem['label']}" + (f" และมี {top_sub_cause['label']} เป็นสาเหตุหลัก" if top_sub_cause else "")
    ci_text = f"{top_ci_item['label']} มี Over SLA สูงสุด ({top_ci_item['over']} Ticket) และควรเป็น Priority ในการตรวจสอบ" if top_ci_item else None
    improvement_focus_text = None
    if top_district and top_problem:
        improvement_focus_text = f"ควรเริ่ม Improve ที่ {top_district['district']} → {top_problem['label']}" + (f" → {top_sub_cause['label']}" if top_sub_cause else "") + (f" → {top_ci_item['label']}" if top_ci_item else "")

    # Improvement Opportunity list - the single best candidate at each
    # level within this province, presented together as "fix these first".
    improvement_opportunities = []
    if top_district and top_district["over"] > 0:
        improvement_opportunities.append({"level": "District", "label": top_district["district"], "over": top_district["over"], "pct_over": top_district["pct_over"], "share_of_total_over": top_district["impact_share"]})
    if top_problem:
        improvement_opportunities.append({"level": "Problem", "label": top_problem["label"], "over": top_problem["over"], "pct_over": top_problem["pct_over"], "share_of_total_over": top_problem["share_of_total_over"]})
    if top_sub_cause:
        improvement_opportunities.append({"level": "Sub-Cause", "label": top_sub_cause["label"], "over": top_sub_cause["over"], "pct_over": top_sub_cause["pct_over"], "share_of_total_over": top_sub_cause["share_of_total_over"]})
    if top_ci_item:
        improvement_opportunities.append({"level": "CI_Name", "label": top_ci_item["label"], "over": top_ci_item["over"], "pct_over": top_ci_item["pct_over"], "share_of_total_over": top_ci_item["share_of_total_over"]})

    return {
        "province": province, "region": region,
        "performance": {
            "total": total, "over": over, "pct_over": pct_over,
            "region_total": region_total, "region_over": region_over, "region_pct_over": region_pct_over,
            "rank": rank, "total_provinces": len(province_ranking["rows"]),
            "is_drag_on_region": is_drag,
        },
        "trend": {"daily": daily_trend, "weekly": weekly_trend},
        "district_table": district_table,
        "district_pareto": district_pareto,
        "root_cause": {"top_problems": top_problems, "top_sub_causes": top_sub_causes, "top_ci": top_ci},
        "pareto": {"district": district_pareto, "problem": problem_pareto, "sub_cause": sub_cause_pareto, "ci": ci_pareto},
        "improvement_opportunities": improvement_opportunities,
        "insight": {
            "key_finding": key_finding, "focus_point": focus_point,
            "root_cause": root_cause_text, "ci_focus": ci_text, "improvement_focus": improvement_focus_text,
        },
    }


def drill_down_province_detail(rows, province, district=None, problem=None, sub_cause=None, top_n=15):
    """Province -> District -> Problem -> Sub-Cause -> CI, one level at a
    time - separate from drill_down_root_cause above (which drills
    Problem->Sub-Cause->CI->Province in that different order for the
    existing Root Cause section) so neither function's behavior changes
    for the other's callers. Always scoped to `province` first."""
    scoped_rows = [r for r in rows if r["province"] == province]
    if district:
        scoped_rows = [r for r in scoped_rows if _normalize_district(r["DISTRICT_EN"]) == district]
    if problem:
        scoped_rows = [r for r in scoped_rows if r["PROBLEM"] == problem]
    if sub_cause:
        scoped_rows = [r for r in scoped_rows if r["SUB_CAUSE"] == sub_cause]
    total_over_in_scope = sum(1 for r in scoped_rows if r["TICKET_SLA"] == "over")

    result = {"matched_over_tickets": total_over_in_scope}
    if not district:
        result["next_level"] = "DISTRICT_EN"
        counts = {}
        for r in scoped_rows:
            d = _normalize_district(r["DISTRICT_EN"])
            counts.setdefault(d, {"total": 0, "over": 0})
            counts[d]["total"] += 1
            if r["TICKET_SLA"] == "over":
                counts[d]["over"] += 1
        ranked = [{"label": k, "total": v["total"], "over": v["over"], "pct_over": _pct(v["over"], v["total"]),
                   "share_of_total_over": _pct(v["over"], total_over_in_scope)} for k, v in counts.items() if v["over"] > 0]
        ranked.sort(key=lambda r: -r["over"])
        result["options"] = ranked[:top_n]
    elif not problem:
        result["next_level"] = "PROBLEM"
        result["options"] = _ranked_group(scoped_rows, "PROBLEM", top_n)
    elif not sub_cause:
        result["next_level"] = "SUB_CAUSE"
        result["options"] = _ranked_group(scoped_rows, "SUB_CAUSE", top_n)
    else:
        result["next_level"] = "CI_Name"
        result["options"] = _ranked_group(scoped_rows, "CI_Name", top_n)
        over_scoped = [r for r in scoped_rows if r["TICKET_SLA"] == "over"]
        result["tickets"] = [
            {"TICKETID": r["TICKETID"], "CI_Name": r["CI_Name"],
             "CREATIONDATE": r["CREATIONDATE"].strftime("%Y-%m-%d %H:%M:%S"),
             "TARGETFINISH": r["TARGETFINISH"].strftime("%Y-%m-%d %H:%M:%S")}
            for r in over_scoped[:200]
        ]
    return result


def build_sla_improvement_response(top_n=15):
    """Orchestrates every analysis above from whatever's currently in the
    in-memory store - raises if no data has been imported yet (the route
    handles that as a normal "please import a file" response, not an
    error)."""
    with _store_lock:
        rows = _store["rows"]
        imported_at = _store["imported_at"]
        filename = _store["filename"]
        import_warnings = _store["warnings"]
    if rows is None:
        return None

    province_ranking = build_province_ranking(rows)  # all-time - still feeds Executive KPI / War Room / Management table below, unchanged
    # Windowed variants purely for the Province Ranking section's own
    # selector (per explicit request) - "all" is the same all-time
    # ranking above, just included here too so the frontend can switch
    # without a second API call.
    province_ranking_windows = {
        "all": province_ranking,
        "3d": build_province_ranking(_filter_last_n_days(rows, 3)),
        "5d": build_province_ranking(_filter_last_n_days(rows, 5)),
        "7d": build_province_ranking(_filter_last_n_days(rows, 7)),
    }
    trend = build_trend(rows)
    improvement_heatmap = build_improvement_heatmap(rows)
    root_cause = build_root_cause(rows, top_n=top_n)
    impact_risk = build_impact_risk(rows)
    executive_kpi = build_executive_kpi(rows, province_ranking, root_cause)
    war_room = build_war_room(rows, root_cause, province_ranking)
    management_table = build_management_table(province_ranking, impact_risk)

    return {
        "imported_at": imported_at, "filename": filename, "import_warnings": import_warnings,
        "row_count": len(rows),
        "executive_kpi": executive_kpi,
        "war_room": war_room,
        "province_ranking": province_ranking,
        "province_ranking_windows": province_ranking_windows,
        "trend": trend,
        "improvement_heatmap": improvement_heatmap,
        "root_cause": root_cause,
        "impact_risk": impact_risk,
        "management_table": management_table,
    }
