"""
pending_ticket.py
------------------
NOR2026 / BBTEC NOC — "Pending Ticket" tab: an operational worksheet where the
team logs what's being done about each pending ticket (problem group, action
team, free-text notes, a work-photo link, a planned close date).

Data flow:
  - Ticket facts (TICKETID, CREATIONDATE, SEVERITY, ...) are read LIVE from
    the same raw sheet realtime_monitor.py uses (REALTIME_SHEET_ID, gid=0).
    That tab is overwritten wholesale by an external export process every
    hour, so nothing the team types can live there - it would vanish on the
    next refresh.
  - Team-entered fields are instead stored in a SEPARATE tab within the SAME
    spreadsheet, "TicketWorkLog", keyed by TICKETID. Every page load joins
    live ticket rows with any saved work-log row for that TICKETID. A ticket
    that's still pending next hour just gets rejoined with the same saved
    work-log data - nothing is lost. A ticket that clears (stops appearing in
    the live sheet) simply stops showing up; its old work-log row is left in
    place (harmless leftover, not actively cleaned up).

Priority (P0/P1/P2): identical formula to realtime_monitor.py.
NANO flag: NN_ClusterID not empty / not "None" -> "NANO".
"""

import logging
import re
import threading
import time
from datetime import datetime, date

from pending_trend import get_drive_and_sheets_clients, bangkok_now, AGING_COLORS, AGING_ORDER, OVER_24H_AGING_KEYS
from mateline_status import build_mateline_status_lookup
from realtime_monitor import REALTIME_SHEET_ID, REALTIME_WORKSHEET_GID, _parse_dt, _classify_priority

log = logging.getLogger(__name__)

ALLOWED_SEVERITIES = {"SA1", "SA2", "SA3", "SA4", "NSA1", "NSA2", "NSA3", "NSA4"}
PENDING_TICKET_REGIONS = {"NOR1", "NOR2"}  # narrower than pending_trend's ALLOWED_REGIONS on purpose

# Sort/group order: named bookmarks first (in this exact order), then
# anything else lumped into "Others".
BOOKMARK_SORT_ORDER = ["7.MB with SA1-4", "4.FBB with SA1-4", "3. All NW Incident NSA1-2"]

LIVE_COLUMNS = [
    "TICKETID", "CREATIONDATE", "TARGETFINISH", "SEVERITY", "SUBJECT", "CINAME",
    "TRUEOWNERGROUP", "Bookmark", "Aging_Flag_Group", "SUBDISTRICT", "DISTRICT",
    "Tech_Team", "Tech_Status", "CLASSIFICATION", "Subimpact",
]

GROUP_PROBLEM_OPTIONS = [
    "Spare part - Spare part not available",
    "Spare part - Wait on site to replace",
    "Budget required - Approve in progress",
    "Site renovate - area is still closed",
    "Site permission - Requesting or pending approval",
    "Site permission - Approved and awaiting site entry",
    "Site dismantle - Wait for configuration removal",
    "Improvement - Awarded subcons",
    "Improvement - Track B",
    "Wait for other parties - Back office/Vendor support",
    "Wait for other parties - Other operators/Provider",
    "Risk Area - Protest/Dangerous area",
    "Disaster - Strom/Flood/Earthquake",
    "Difficulte work - Site equipment was stolen",
    "Difficulte work - Work at high",
    "Fiber optic - Fire burn/High voltage short circuit",
    "Fiber optic \u2013 Accident/Constuction",
    "Fiber optic - PEA/MEA Maintenance transmission line",
    "Fiber optic - High loss/Degrade/Animal",
    "Workload - Assigning team access site",
    "Solar Cell - Off Grid",
    "Clear",
]

ACTION_TEAM_OPTIONS = ["OFC", "NODE", "SUP", "ENG Zone", "Special Team", "Planwork", "Track 7", "E//", "Other Vendor"]

WORK_LOG_SHEET = "TicketWorkLog"
WORK_LOG_HEADER = [
    "ticket_id", "group_problem", "action_team", "detail", "image_link",
    "plan_closed_date", "updated_at", "updated_by",
]

# External mirror: every build of the Pending Ticket table also gets written
# here in full, exactly matching what's shown on the web page, plus an
# insert_time column stamping when that write happened.
EXPORT_SHEET_ID = "10Y3Pyp6-MqlrcxiDXdAjTNWWZLgDWp94dZvU51lOxWU"
EXPORT_WORKSHEET_GID = 1260767128
EXPORT_HEADER = [
    "TICKETID", "CINAME", "SUBJECT", "Priority", "CREATIONDATE", "TARGETFINISH",
    "SEVERITY", "TRUEOWNERGROUP", "Bookmark", "Aging_Flag_Group", "SUBDISTRICT",
    "DISTRICT", "Tech_Team", "Tech_Status", "NANO", "Group_Problem", "Action_Team",
    "Detail", "Image_Link", "Plan_Closed_Date", "Updated_At", "Updated_By", "insert_time",
]


# ---------------------------------------------------------------------------
# Live ticket rows
# ---------------------------------------------------------------------------

def _get_worksheet(gs_client):
    sh = gs_client.open_by_key(REALTIME_SHEET_ID)
    for ws in sh.worksheets():
        if ws.id == REALTIME_WORKSHEET_GID:
            return ws
    return sh.sheet1


# Two very different caching needs share this file:
#
# - live ticket rows: fed by an EXTERNAL system (SCCD+ITSM) that re-stamps
#   `insert_time` on its own frequent cadence (every few minutes, as new
#   tickets stream in) - not the rarely-changing kind of value the
#   insert_time trick works well for. An earlier version of this function
#   read that cell first as a "cheap check" before deciding whether to do
#   the full get_all_records() read, mirroring realtime_monitor.py - but
#   because insert_time here changes on nearly every load anyway, that
#   turned one Sheets API round trip into two almost every time (the cheap
#   check, then the full read right after since insert_time had already
#   moved on) - net SLOWER, which is what caused the "หน้านี้มันนานไป"
#   report. Reverted to a plain short-TTL cache below: no extra round trip,
#   just reuse the same read for anyone loading within the TTL window.
# - work log: written entirely through save_work_log_entry() in THIS app, so
#   it explicitly invalidates its cache on every save - safe to cache for a
#   long time since staleness can only come from someone editing the sheet
#   directly outside the app.
#
# Both caches are safe to rely on only because Gunicorn runs a single worker
# process (see Procfile) - this in-memory cache is shared by every request.
# Raising the worker count would split it across processes and let stale
# reads slip through; use more threads instead if more concurrency is ever
# needed, not more workers.
LIVE_ROWS_CACHE_TTL_SECONDS = 45   # this sheet is shared by nearly every feature on the dashboard now
                                    # (Pending Ticket, P0 Only, Flood NAN, P0 snapshot comparison's "current"
                                    # count...) - too short a TTL here means cumulative Sheets API reads
                                    # across everyone's tabs adds up fast and trips the per-minute read quota.
                                    # 45s keeps data fresh enough for NOC monitoring while cutting call volume ~3x.
WORK_LOG_CACHE_TTL_SECONDS = 7200   # 2h; save_work_log_entry invalidates explicitly
_live_rows_cache = {"data": None, "ts": 0}
_work_log_cache = {"data": None, "ts": 0}
_work_log_row_index_cache = {}  # ticket_id -> sheet row number, kept alongside _work_log_cache
_cache_lock = threading.Lock()


def fetch_live_rows(gs_client, use_cache=True):
    now = time.monotonic()
    if use_cache:
        with _cache_lock:
            if _live_rows_cache["data"] is not None and (now - _live_rows_cache["ts"]) < LIVE_ROWS_CACHE_TTL_SECONDS:
                return _live_rows_cache["data"]
            # Cache is cold/stale - fetch WHILE STILL HOLDING the lock, so a
            # second thread arriving concurrently blocks here instead of
            # also deciding independently that a fresh Sheets read is
            # needed. Without this, N callers hitting a cold cache at once
            # (e.g. a page that reads this sheet twice in one request, or
            # two people loading the same tab within the same instant) each
            # do their own separate read instead of sharing one - a classic
            # cache-stampede, and the likely cause of a burst of Sheets API
            # reads tripping the per-minute read quota.
            ws = _get_worksheet(gs_client)
            rows = ws.get_all_records()
            _live_rows_cache["data"] = rows
            _live_rows_cache["ts"] = time.monotonic()
            return rows

    ws = _get_worksheet(gs_client)
    return ws.get_all_records()



def _bookmark_sort_key(bookmark):
    try:
        return BOOKMARK_SORT_ORDER.index(bookmark)
    except ValueError:
        return len(BOOKMARK_SORT_ORDER)  # "Others" - sorts after the named ones


# ---------------------------------------------------------------------------
# Work-log persistence (separate tab, survives the live sheet's hourly wipe)
# ---------------------------------------------------------------------------

def _ensure_work_log_tab(spreadsheet):
    try:
        return spreadsheet.worksheet(WORK_LOG_SHEET)
    except Exception:
        ws = spreadsheet.add_worksheet(title=WORK_LOG_SHEET, rows=2000, cols=len(WORK_LOG_HEADER))
        ws.append_row(WORK_LOG_HEADER)
        return ws


def load_work_log(gs_client, use_cache=True):
    """Returns {ticket_id: {group_problem, action_team, detail, image_link,
    plan_closed_date, updated_at, updated_by}} for every saved row. Also
    populates _work_log_row_index_cache (ticket_id -> sheet row number) as
    a side effect, so save_work_log_entry can look up whether a ticket
    already has a row without its own separate read."""
    now = time.monotonic()
    if use_cache:
        with _cache_lock:
            if _work_log_cache["data"] is not None and (now - _work_log_cache["ts"]) < WORK_LOG_CACHE_TTL_SECONDS:
                return _work_log_cache["data"]
    sh = gs_client.open_by_key(REALTIME_SHEET_ID)
    ws = _ensure_work_log_tab(sh)
    records = ws.get_all_values()[1:]  # skip header
    out = {}
    row_index = {}
    for i, row in enumerate(records):
        if not row or not row[0]:
            continue
        padded = row + [""] * (len(WORK_LOG_HEADER) - len(row))
        out[padded[0]] = {
            "group_problem": padded[1], "action_team": padded[2], "detail": padded[3],
            "image_link": padded[4], "plan_closed_date": padded[5],
            "updated_at": padded[6], "updated_by": padded[7],
        }
        row_index[padded[0]] = i + 2  # +1 for header, +1 for 1-based sheet rows
    if use_cache:
        with _cache_lock:
            _work_log_cache["data"] = out
            _work_log_cache["ts"] = now
            _work_log_row_index_cache.clear()
            _work_log_row_index_cache.update(row_index)
    return out


def _invalidate_work_log_cache():
    with _cache_lock:
        _work_log_cache["data"] = None
        _work_log_cache["ts"] = 0


# Ordered keyword rules for _auto_categorize_subject - derived from and
# validated against a 612-ticket labeled reference set (100% match), all
# NSA3/NSA4 tickets. Order matters: earlier rules win when a subject
# matches more than one (e.g. "Service Degraded | RRU-5 HW Partial Fault"
# needs HW Partial Fault checked before the generic Radio Performance
# Degraded catch-all, or it'd never be reached).
SUBJECT_CATEGORY_RULES = [
    ("Cell Up/Down Alarm", ["cell up/down"]),
    ("Intrusion Alarm (ความปลอดภัยสถานี)", ["intrusion", "relay_alarm_major_open_doo", "open_door", "open door", "(ason "]),
    ("Link Failure/Degraded", ["link failure", "link degraded", "ethernet link", "link_failure", "link failure up/down"]),
    ("HW Partial Fault", ["hw partial fault"]),
    ("IPRAN Hardware Alarm (Fan/Board)", ["ipran alarm board", "ipran alarm  board"]),
    ("Environmental Alarm (Temperature/Smoke)", ["temperature"]),
    ("Radio Performance Degraded", ["service degraded", "service unavailable", "resource allocation failure",
        "sleeping cell", "increased ber", "internal interference", "performance degraded",
        "ออกสลับกัน", "node group sync loss", "cell sleep"]),
    ("External Alarm - Power/Site Infra", ["external alarm", "input power failure", "lossofmains", "main ac power failure", "rectifier"]),
    ("VSWR Alarm (สายอากาศ/ฟีดเดอร์)", ["vswr", "rf reflected power high"]),
    ("Solar/Inverter Power Alarm", ["solar cell", "inverter"]),
    ("Antenna/RET Alarm", ["rxdiversitylost", "retportcurrenttoohigh", "antennasystemproblem",
        "digitalcable_cablefailure", "antennabranch", "retdevice_retfailure", "rx branch imbalance",
        "communicationlostwithret", "antenna calibration", "current too high"]),
    ("Equipment Fault (RET/Connection)", ["rap no connection", "ret not calibrated", "ret failure", "no_connection"]),
    ("Sync/Timing Alarm", ["timesyncio", "sync ptp", "time reachability"]),
    ("IPRAN Power/Optical Alarm", ["ipran"]),
    ("Fan Alarm", ["fan fail", "fan failure"]),
    ("RF Module Failure", ["rf module failure"]),
    ("Battery/Power Management Alarm", ["paco-casa", "energy_cell", "swap battery"]),
    ("DWDM/Optical Transport Alarm", ["dwdm", "mossman input power"]),
    ("Microwave Link Alarm", ["microwave"]),
    ("Complaint / No Traffic", ["nbtc"]),
    ("CPRI/RRU Fiber Alarm", ["cpri", "fan rru"]),
    ("Transmission Power Alarm", ["input_b power low", "input power low"]),
    ("FTTX/OLT Alarm", ["olt", "ftth", "fiberhome", "edfa"]),
    ("Manual Request/Coordination (แจ้งประสานงาน)", ["request check", "เข้าแก้ไขปัญหา", "dpo request", "optimiz", "support for check"]),
    ("Hardware/Software Fault (FRU/HW/SW)", ["install sw", "rejectsignalfromhardware", "fan speed continuously",
        "hw fault", "swerror", "sw error", "resource activation timeout", "license key not available",
        "lan switch abnormality", "resource configuration failure", "fru general problem",
        "mo configuration not consistent"]),
]
SUBJECT_CATEGORY_OTHER = "Other/Uncategorized"


def _auto_categorize_subject(subject):
    """Groups a ticket SUBJECT line into one of the fixed categories in
    SUBJECT_CATEGORY_RULES via ordered keyword matching (see that list's
    docstring for how it was derived/validated). Falls back to
    SUBJECT_CATEGORY_OTHER for anything that matches none of them."""
    s = str(subject or "").lower()
    if not s.strip():
        return SUBJECT_CATEGORY_OTHER
    for category, keywords in SUBJECT_CATEGORY_RULES:
        if any(k in s for k in keywords):
            return category
    return SUBJECT_CATEGORY_OTHER


def _row_number_from_append_response(resp):
    """Extracts the row number Sheets actually appended to from
    append_row()'s raw API response (updates.updatedRange, e.g.
    'TicketWorkLog!A15:H15') - avoids a follow-up read to find out where
    the new row landed."""
    try:
        updated_range = resp["updates"]["updatedRange"]
        cell_ref = updated_range.split("!")[1].split(":")[0]  # "A15"
        return int(re.search(r"\d+", cell_ref).group())
    except (KeyError, IndexError, AttributeError, ValueError):
        return None


def save_work_log_entry(gs_client, ticket_id, fields, updated_by=None):
    """fields: dict with any of group_problem/action_team/detail/image_link/
    plan_closed_date. Upserts the row for ticket_id, stamping updated_at (and
    updated_by once a login system exists - for now defaults to 'unknown').

    Looks up whether ticket_id already has a row via the cached row-index
    (populated by load_work_log(), which this calls first to guarantee it's
    warm) instead of its own ws.col_values(1) read - that used to run
    uncached on every single save, which under normal multi-person use
    (several saves in quick succession while working an incident) was
    enough on its own to trip the Sheets API's per-minute read quota.
    Updates the cache in place afterward rather than invalidating it, so a
    burst of consecutive saves stays fully cache-driven - the sheet is only
    ever actually re-read after WORK_LOG_CACHE_TTL_SECONDS of no saves, or
    if something edited the sheet directly outside the app."""
    load_work_log(gs_client)  # ensures _work_log_row_index_cache is warm (cache hit unless >2h idle)
    sh = gs_client.open_by_key(REALTIME_SHEET_ID)
    ws = _ensure_work_log_tab(sh)

    now_str = bangkok_now().strftime("%Y-%m-%d %H:%M:%S")
    row_values = [
        ticket_id,
        fields.get("group_problem", ""),
        fields.get("action_team", ""),
        fields.get("detail", ""),
        fields.get("image_link", ""),
        fields.get("plan_closed_date", ""),
        now_str,
        updated_by or "unknown",
    ]

    with _cache_lock:
        row_idx = _work_log_row_index_cache.get(ticket_id)

    if row_idx is not None:
        ws.update(f"A{row_idx}:H{row_idx}", [row_values])
    else:
        resp = ws.append_row(row_values)
        row_idx = _row_number_from_append_response(resp)
        if row_idx is None:
            # Couldn't tell where the row landed - don't risk caching a wrong
            # index (a future save for this ticket would silently append a
            # duplicate row instead of updating). Force a full re-read next time.
            _invalidate_work_log_cache()
            return row_values

    with _cache_lock:
        if _work_log_cache["data"] is not None:
            _work_log_cache["data"][ticket_id] = {
                "group_problem": row_values[1], "action_team": row_values[2], "detail": row_values[3],
                "image_link": row_values[4], "plan_closed_date": row_values[5],
                "updated_at": row_values[6], "updated_by": row_values[7],
            }
        _work_log_row_index_cache[ticket_id] = row_idx
    return row_values


def rename_group_problem_value(gs_client, old_value, new_value):
    """One-time cleanup: renames every occurrence of `old_value` in the
    group_problem column of TicketWorkLog to `new_value` (e.g. consolidating
    two similar options into one after the dropdown list changes). Returns
    how many rows were updated."""
    sh = gs_client.open_by_key(REALTIME_SHEET_ID)
    ws = _ensure_work_log_tab(sh)
    values = ws.get_all_values()
    if not values:
        return 0
    updates = []
    changed = 0
    for i, row in enumerate(values[1:], start=2):  # skip header
        if len(row) > 1 and row[1] == old_value:
            updates.append({"range": f"B{i}", "values": [[new_value]]})
            changed += 1
    if updates:
        ws.batch_update(updates, value_input_option="RAW")
        _invalidate_work_log_cache()
    return changed


def _get_export_worksheet(gs_client):
    sh = gs_client.open_by_key(EXPORT_SHEET_ID)
    for ws in sh.worksheets():
        if ws.id == EXPORT_WORKSHEET_GID:
            return ws
    return sh.sheet1


def _ticket_to_export_row(t, insert_time_str=""):
    """Builds one EXPORT_HEADER-shaped row from a ticket entry dict - shared
    by the external mirror export and the on-demand Excel/Google Sheet
    export buttons, so all three always produce the same column layout."""
    return [
        t.get("TICKETID", ""), t.get("CINAME", ""), t.get("SUBJECT", ""), t.get("priority", ""),
        t.get("CREATIONDATE", ""), t.get("TARGETFINISH", ""), t.get("SEVERITY", ""),
        t.get("TRUEOWNERGROUP", ""), t.get("Bookmark", ""), t.get("Aging_Flag_Group", ""),
        t.get("SUBDISTRICT", ""), t.get("DISTRICT", ""), t.get("Tech_Team", ""), t.get("Tech_Status", ""),
        t.get("nano", ""), t.get("group_problem", ""), t.get("action_team", ""), t.get("detail", ""),
        t.get("image_link", ""), t.get("plan_closed_date", ""), t.get("updated_at", ""), t.get("updated_by", ""),
        insert_time_str,
    ]


def export_to_external_sheet(gs_client, tickets, insert_time_str):
    """Mirrors the full current Pending Ticket table (exactly what's on the
    web page) into the external sheet, overwriting whatever was there before.
    Called once per build_pending_ticket_response(), so it stays in sync
    every time the tab is loaded or refreshed."""
    ws = _get_export_worksheet(gs_client)
    rows = [_ticket_to_export_row(t, insert_time_str) for t in tickets]
    ws.clear()
    ws.update("A1", [EXPORT_HEADER] + rows, value_input_option="RAW")
    log.info("Exported %d rows to external Pending Ticket mirror sheet", len(rows))


# ---------------------------------------------------------------------------
# Combined response
# ---------------------------------------------------------------------------

def _multi_filter(entries, key, values):
    if not values:
        return entries
    values_set = set(values)
    return [e for e in entries if str(e.get(key, "")).strip() in values_set]


def _fetch_full_ticket_entries(gs_client):
    """Builds the full, export-ready ticket entry list (every field
    EXPORT_HEADER needs) - shared by build_pending_ticket_response and
    the background-export trigger, so exporting doesn't depend on which
    page's response function happens to call it."""
    all_rows = fetch_live_rows(gs_client)
    now_dt = bangkok_now()
    today = now_dt.date()

    scoped = [
        r for r in all_rows
        if str(r.get("Region", "")).strip() in PENDING_TICKET_REGIONS
        and str(r.get("SEVERITY", "")).strip() in ALLOWED_SEVERITIES
    ]

    work_log = load_work_log(gs_client)

    def build_entry(r):
        ticket_id = str(r.get("TICKETID", "")).strip()
        nn_cluster = r.get("NN_ClusterID")
        nano = "NANO" if nn_cluster not in (None, "", "None") else ""
        over_sla_day = r.get("Over_SLA_Day")
        try:
            over_sla_day = float(over_sla_day)
        except (TypeError, ValueError):
            over_sla_day = 0

        entry = {c: r.get(c) for c in LIVE_COLUMNS}
        entry["over_sla_day"] = over_sla_day
        entry["priority"] = _classify_priority(r.get("TARGETFINISH"), now_dt)
        entry["nano"] = nano
        entry["insert_time"] = r.get("insert_time", "")
        wl = work_log.get(ticket_id, {})
        entry["group_problem"] = wl.get("group_problem", "")
        entry["action_team"] = wl.get("action_team", "")
        entry["detail"] = wl.get("detail", "")
        entry["image_link"] = wl.get("image_link", "")
        entry["plan_closed_date"] = wl.get("plan_closed_date", "")
        entry["updated_at"] = wl.get("updated_at", "")
        entry["updated_by"] = wl.get("updated_by", "")
        entry["plan_closed_overdue"] = False
        if entry["plan_closed_date"]:
            try:
                pcd = datetime.strptime(entry["plan_closed_date"], "%Y-%m-%d").date()
                entry["plan_closed_overdue"] = pcd < today
            except ValueError:
                pass
        return entry

    return [build_entry(r) for r in scoped]


def trigger_background_export(gs_client, all_entries=None):
    """Fire-and-forget mirror export to the external tracking sheet. Called
    from EVERY page that shows live ticket data (Pending Ticket, P0 Only) -
    not just Pending Ticket - so the mirror sheet stays fresh no matter
    which page people are actually working from. Safe to call often: this
    is exactly what was silently going stale before, because only Pending
    Ticket's own response builder used to trigger it."""
    if all_entries is None:
        all_entries = _fetch_full_ticket_entries(gs_client)
    export_insert_time = bangkok_now().strftime("%Y-%m-%d %H:%M:%S")

    def _export_in_background(entries, insert_time_str):
        try:
            export_to_external_sheet(gs_client, entries, insert_time_str)
        except Exception:
            log.exception("Failed to export Pending Ticket table to external mirror sheet - continuing anyway")

    threading.Thread(target=_export_in_background, args=(all_entries, export_insert_time), daemon=True).start()
    return export_insert_time


def build_pending_ticket_xlsx(matched_entries):
    """Builds an in-memory .xlsx workbook of the given ticket entries (same
    EXPORT_HEADER columns as the mirror sheet), for the on-demand Excel
    export button - not written to disk, returned as bytes for a Flask
    file download response."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Pending Ticket"
    ws.append(EXPORT_HEADER)
    header_fill = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    rows = [_ticket_to_export_row(t) for t in matched_entries]
    for row in rows:
        ws.append(row)
    for i, header in enumerate(EXPORT_HEADER, start=1):
        col_letter = get_column_letter(i)
        max_len = max([len(str(header))] + [len(str(row[i - 1])) for row in rows] or [10])
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_pending_ticket_response(gs_client=None, bookmark_filter=None, trueowner_filter=None,
                                   severity_filter=None, district_filter=None, group_problem_filter=None,
                                   aging_filter=None):
    """bookmark_filter/trueowner_filter/severity_filter/district_filter/group_problem_filter/
    aging_filter: each an optional LIST of values (multi-select) - None or empty means "no
    filter, include everything in that dimension"."""
    if gs_client is None:
        _, gs_client = get_drive_and_sheets_clients()

    all_entries = _fetch_full_ticket_entries(gs_client)
    all_entries.sort(key=lambda t: (_bookmark_sort_key(str(t.get("Bookmark", "")).strip()), -t["over_sla_day"]))

    present_agings = {str(e.get("Aging_Flag_Group", "")).strip() for e in all_entries if e.get("Aging_Flag_Group")}
    filter_options = {
        "bookmarks": sorted({str(e.get("Bookmark", "")).strip() for e in all_entries if e.get("Bookmark")}),
        "trueowners": sorted({str(e.get("TRUEOWNERGROUP", "")).strip() for e in all_entries if e.get("TRUEOWNERGROUP")}),
        "severities": sorted({str(e.get("SEVERITY", "")).strip() for e in all_entries if e.get("SEVERITY")}),
        # Kept in canonical Aging_Flag_Group order (not alphabetical) so the
        # dropdown reads 1)->6) like everywhere else Aging is shown.
        "agings": [a for a in AGING_ORDER if a in present_agings],
    }

    # District counts are faceted on the OTHER active filters (bookmark/trueowner/
    # severity/aging) so the numbers next to each district option stay accurate
    # as those filters change - just not on the district filter itself.
    pre_district = _multi_filter(all_entries, "Bookmark", bookmark_filter)
    pre_district = _multi_filter(pre_district, "TRUEOWNERGROUP", trueowner_filter)
    pre_district = _multi_filter(pre_district, "SEVERITY", severity_filter)
    pre_district = _multi_filter(pre_district, "Aging_Flag_Group", aging_filter)
    district_counts = {}
    for e in pre_district:
        d = str(e.get("DISTRICT", "")).strip()
        if d:
            district_counts[d] = district_counts.get(d, 0) + 1
    filter_options["districts"] = [
        {"district": d, "count": c}
        for d, c in sorted(district_counts.items(), key=lambda kv: kv[1], reverse=True)
    ]

    matched_entries = _multi_filter(pre_district, "DISTRICT", district_filter)
    matched_entries = _multi_filter(matched_entries, "group_problem", group_problem_filter)

    filter_options["group_problems"] = sorted({
        str(e.get("group_problem", "")).strip() for e in all_entries if e.get("group_problem")
    })

    # Mirrors the full table to an external tracking sheet - shared trigger
    # used by every page that shows live ticket data, not just this one.
    export_insert_time = trigger_background_export(gs_client, all_entries)

    # fetch_live_rows is cached (short TTL), so this is cheap - just need
    # the raw sheet's own insert_time stamp (not part of LIVE_COLUMNS) to
    # show how fresh the SOURCE data is, separate from export_insert_time
    # (when THIS export just ran).
    all_rows = fetch_live_rows(gs_client)

    return {
        "total": len(matched_entries),
        "filter_options": filter_options,
        "group_problem_options": GROUP_PROBLEM_OPTIONS,
        "action_team_options": ACTION_TEAM_OPTIONS,
        "aging_colors": AGING_COLORS,
        "insert_time": all_rows[0].get("insert_time") if all_rows else None,
        "export_insert_time": export_insert_time,
        "tickets": matched_entries,
    }


# ---------------------------------------------------------------------------
# Exclusive Pending — morning-meeting summary: for each Bookmark, a
# Group Problem x Aging_Flag_Group matrix, plus the ticket-level detail for
# anything sitting in the four "over SLA" aging buckets (1-4; excludes
# "< 1 day" and "Within SLA" since those aren't meeting-worthy yet).
# ---------------------------------------------------------------------------

EXCLUSIVE_BOOKMARK_ORDER = BOOKMARK_SORT_ORDER + ["NSA3-4"]
UNSPECIFIED_GROUP_PROBLEM = "(ยังไม่ระบุ)"
UNSPECIFIED_AGING = "(ไม่ระบุ)"


def _exclusive_bookmark_label(raw):
    # Catch-all bucket for anything not in the 3 named Bookmark categories -
    # in practice this is always NSA3/NSA4 tickets, so the label reflects
    # that directly instead of the generic "Others" (single source of
    # truth: this flows through to every page - KPI cards, section
    # headers, dropdowns, LINE summaries - without needing separate edits).
    raw = str(raw or "").strip()
    return raw if raw in BOOKMARK_SORT_ORDER else "NSA3-4"


def build_online_remaining_hours_by_province(entries):
    """For the "4.FBB with SA1-4" (Online) bookmark specifically - groups
    by Province, and for each: bucket counts by remaining time to
    TARGETFINISH (already overdue / <=2h / <=6h / <=12h / >12h), plus the
    single most urgent ticket (soonest TARGETFINISH) as a concrete
    reference. Sorted so the province with the most urgent situation
    (soonest TARGETFINISH) appears first - "เรียงจากน้อยไปมาก" per
    explicit request - so a reader can see at a glance where to focus
    first for planning. Entries with no parseable TARGETFINISH are
    excluded (their remaining time genuinely isn't known, not guessed)."""
    online_entries = [e for e in entries if e["Bookmark"] == "4.FBB with SA1-4" and e["remaining_hours"] is not None]
    by_province = {}
    for e in online_entries:
        by_province.setdefault(e["PROVINCE"], []).append(e)

    rows = []
    for prov, prov_entries in by_province.items():
        prov_entries = sorted(prov_entries, key=lambda e: e["remaining_hours"])
        most_urgent = prov_entries[0]
        buckets = {"over": 0, "le_2h": 0, "le_6h": 0, "le_12h": 0, "gt_12h": 0}
        for e in prov_entries:
            h = e["remaining_hours"]
            if h < 0:
                buckets["over"] += 1
            elif h <= 2:
                buckets["le_2h"] += 1
            elif h <= 6:
                buckets["le_6h"] += 1
            elif h <= 12:
                buckets["le_12h"] += 1
            else:
                buckets["gt_12h"] += 1
        rows.append({
            "region": prov_entries[0]["Region"], "province": prov, "total": len(prov_entries),
            **buckets,
            "most_urgent_ticket": most_urgent["TICKETID"],
            "most_urgent_remaining_hours": most_urgent["remaining_hours"],
            "most_urgent_subject": most_urgent["SUBJECT"],
        })
    rows.sort(key=lambda r: r["most_urgent_remaining_hours"])
    return rows


def build_exclusive_pending_response(gs_client=None, priority_filter=None, restrict_to_over_sla=True):
    """Morning-meeting view: how many pending tickets per Bookmark are stuck
    on which Group Problem, broken out by aging bucket - plus the full
    ticket-level list for anything already over SLA (aging buckets 1-4),
    so a team lead can answer CNO questions ticket-by-ticket without
    switching tabs. `priority_filter` (e.g. "P0") restricts EVERYTHING -
    matrix, province breakdowns, plan-date matrix, detail lists - to just
    that priority, for a page like "P0 Only" that needs the exact same
    view scoped down to the most urgent tickets. `restrict_to_over_sla`
    controls the DETAIL list only: True (default, normal Exclusive
    Pending) limits it to the four over-SLA aging buckets; False includes
    every ticket regardless of aging bucket - used by "P0 Only" so it
    shows literally every P0 ticket, not just the ones already overdue."""
    if gs_client is None:
        _, gs_client = get_drive_and_sheets_clients()

    # Also mirrors the full ticket table to the external tracking sheet -
    # same trigger Pending Ticket uses. Needed here too: P0 Only is where
    # the team actually spends most of their time now, so if only Pending
    # Ticket's own load triggered this, the mirror sheet would silently go
    # stale on days nobody happens to open that specific tab.
    trigger_background_export(gs_client)

    all_rows = fetch_live_rows(gs_client)
    now_dt = bangkok_now()
    scoped = [
        r for r in all_rows
        if str(r.get("Region", "")).strip() in PENDING_TICKET_REGIONS
        and str(r.get("SEVERITY", "")).strip() in ALLOWED_SEVERITIES
    ]
    work_log = load_work_log(gs_client)

    # GGS Daily mateline status (external field-team tracking sheet) - a
    # failure here (e.g. the sheet isn't shared with this service account)
    # shouldn't take down the whole page, so every entry just falls back
    # to "(ไม่พบใน MatelineX)" if the lookup couldn't be built at all.
    try:
        mateline_lookup = build_mateline_status_lookup(gs_client, now_dt.strftime("%Y-%m-%d"))
    except Exception:
        log.exception("GGS Daily mateline lookup failed - falling back to empty for this response")
        mateline_lookup = {}

    entries = []
    for r in scoped:
        ticket_id = str(r.get("TICKETID", "")).strip()
        wl = work_log.get(ticket_id, {})
        mateline = mateline_lookup.get(ticket_id.upper()) or {
            "status_mateline": "(ไม่พบใน MatelineX)", "mateline_wo_status": "",
        }
        over_sla_day = r.get("Over_SLA_Day")
        try:
            over_sla_day = float(over_sla_day)
        except (TypeError, ValueError):
            over_sla_day = 0
        bookmark_label = _exclusive_bookmark_label(r.get("Bookmark"))
        priority = _classify_priority(r.get("TARGETFINISH"), now_dt)
        if priority_filter:
            # Explicit exception: on the P0-scoped view (P0 Only), the
            # "Online" bookmark (4.FBB with SA1-4) specifically shows P0
            # AND P1 together - every other bookmark still shows exactly
            # priority_filter, unchanged.
            if priority_filter == "P0" and bookmark_label == "4.FBB with SA1-4":
                allowed_priorities = {"P0", "P1"}
            else:
                allowed_priorities = {priority_filter}
            if priority not in allowed_priorities:
                continue
        target_dt = _parse_dt(r.get("TARGETFINISH"))
        remaining_hours = round((target_dt - now_dt).total_seconds() / 3600, 2) if target_dt else None
        entries.append({
            "TICKETID": ticket_id,
            "SUBJECT": r.get("SUBJECT", ""),
            "subject_category": _auto_categorize_subject(r.get("SUBJECT", "")),
            "CINAME": r.get("CINAME", ""),
            "DISTRICT": r.get("DISTRICT", ""),
            "PROVINCE": r.get("PROVINCE", ""),
            "Region": r.get("Region", ""),
            "TRUEOWNERGROUP": r.get("TRUEOWNERGROUP", ""),
            "priority": priority,
            "Bookmark": bookmark_label,
            "Aging_Flag_Group": str(r.get("Aging_Flag_Group", "")).strip() or UNSPECIFIED_AGING,
            "group_problem": wl.get("group_problem") or UNSPECIFIED_GROUP_PROBLEM,
            "action_team": wl.get("action_team", ""),
            "detail": wl.get("detail", ""),
            "image_link": wl.get("image_link", ""),  # preserved so saving from this page never blanks it out
            "plan_closed_date": wl.get("plan_closed_date", ""),
            "over_sla_day": over_sla_day,
            "status_mateline": mateline["status_mateline"],
            "mateline_wo_status": mateline["mateline_wo_status"],
            "TARGETFINISH": r.get("TARGETFINISH", ""),
            "remaining_hours": remaining_hours,  # None if TARGETFINISH doesn't parse - never guessed
        })

    # Summary matrix: bookmark -> group_problem -> aging_key -> count
    summary = {}
    for e in entries:
        (summary.setdefault(e["Bookmark"], {})
                 .setdefault(e["group_problem"], {})
                 .setdefault(e["Aging_Flag_Group"], 0))
        summary[e["Bookmark"]][e["group_problem"]][e["Aging_Flag_Group"]] += 1

    summary_out = []
    for bm in EXCLUSIVE_BOOKMARK_ORDER:
        if bm not in summary:
            continue
        rows = []
        for gp, counts in summary[bm].items():
            # "over_total" normally means just the over-SLA buckets (1-4) -
            # but when restrict_to_over_sla is off (P0 Only), it should
            # count every bucket instead, so this number and the detail
            # list below (and the KPI breakdown card) all agree on the
            # same total instead of the KPI card silently under-counting.
            if restrict_to_over_sla:
                over_total = sum(counts.get(k, 0) for k in OVER_24H_AGING_KEYS)
            else:
                over_total = sum(counts.values())
            row = {"group_problem": gp, "over_total": over_total, "counts": {
                ag: counts.get(ag, 0) for ag in AGING_ORDER
            }}
            rows.append(row)
        rows.sort(key=lambda r: -r["over_total"])
        bookmark_total = sum(r["over_total"] for r in rows)
        summary_out.append({"bookmark": bm, "rows": rows, "over_total": bookmark_total})

    # Province x Aging_Flag_Group matrix - same shape/logic as the Group
    # Problem matrix above, just grouped by PROVINCE instead. Built from
    # the same full `entries` set (every aging bucket), independent of
    # restrict_to_over_sla.
    province_matrix = {}
    for e in entries:
        prov = str(e["PROVINCE"]).strip() or "(ไม่ระบุจังหวัด)"
        (province_matrix.setdefault(e["Bookmark"], {})
                         .setdefault(prov, {})
                         .setdefault(e["Aging_Flag_Group"], 0))
        province_matrix[e["Bookmark"]][prov][e["Aging_Flag_Group"]] += 1

    province_summary_out = []
    for bm in EXCLUSIVE_BOOKMARK_ORDER:
        if bm not in province_matrix:
            continue
        rows = []
        for prov, counts in province_matrix[bm].items():
            if restrict_to_over_sla:
                over_total = sum(counts.get(k, 0) for k in OVER_24H_AGING_KEYS)
            else:
                over_total = sum(counts.values())
            rows.append({"province": prov, "over_total": over_total, "counts": {
                ag: counts.get(ag, 0) for ag in AGING_ORDER
            }})
        rows.sort(key=lambda r: -r["over_total"])
        bookmark_total = sum(r["over_total"] for r in rows)
        province_summary_out.append({"bookmark": bm, "rows": rows, "over_total": bookmark_total})

    # Detail: over-SLA aging buckets (1-4) only by default, matching the
    # normal Exclusive Pending view - unless restrict_to_over_sla is off
    # (P0 Only), in which case every entry that made it this far (already
    # priority-filtered above) is included regardless of aging bucket.
    detail_entries = [e for e in entries if not restrict_to_over_sla or e["Aging_Flag_Group"] in OVER_24H_AGING_KEYS]
    detail_entries.sort(key=lambda e: -e["over_sla_day"])

    detail_out = []
    unspecified_by_province_out = []
    group_problem_by_plan_date_out = []
    NO_PLAN_DATE = "(ยังไม่วางแผน)"
    for bm in EXCLUSIVE_BOOKMARK_ORDER:
        tickets = [e for e in detail_entries if e["Bookmark"] == bm]
        if not tickets:
            continue
        block = {"bookmark": bm, "tickets": tickets, "total": len(tickets)}

        # Category x Aging_Flag_Group matrix - only for the NSA3-4 group,
        # per request (that bucket's tickets are varied enough, and few
        # enough people are triaging it by hand, that grouping by
        # recurring problem pattern actually helps there specifically).
        # Same shape as the other matrices on this page (Group Problem x
        # Aging, Province x Aging), so which category is piling up in
        # which aging bucket is visible at a glance.
        if bm == "NSA3-4":
            category_aging_matrix = {}
            for e in tickets:
                cat = e["subject_category"]
                ag = e["Aging_Flag_Group"]
                category_aging_matrix.setdefault(cat, {}).setdefault(ag, 0)
                category_aging_matrix[cat][ag] += 1
            category_aging_rows = []
            for cat, counts in category_aging_matrix.items():
                row_total = sum(counts.values())
                category_aging_rows.append({
                    "category": cat, "over_total": row_total,
                    "counts": {ag: counts.get(ag, 0) for ag in AGING_ORDER},
                })
            category_aging_rows.sort(key=lambda r: -r["over_total"])
            block["subject_category_by_aging"] = category_aging_rows

            # Category x Province matrix - same rows (Category), different
            # column dimension, for the same table's "switch to Province"
            # toggle (columns swap, rows stay Category either way).
            category_province_matrix = {}
            provinces_seen = set()
            for e in tickets:
                cat = e["subject_category"]
                prov = (e.get("PROVINCE") or "").strip() or "(ไม่ระบุจังหวัด)"
                provinces_seen.add(prov)
                category_province_matrix.setdefault(cat, {}).setdefault(prov, 0)
                category_province_matrix[cat][prov] += 1
            province_columns = sorted(provinces_seen)
            category_province_rows = []
            for cat, counts in category_province_matrix.items():
                row_total = sum(counts.values())
                category_province_rows.append({
                    "category": cat, "over_total": row_total,
                    "counts": {p: counts.get(p, 0) for p in province_columns},
                })
            category_province_rows.sort(key=lambda r: -r["over_total"])
            block["subject_category_by_province"] = {
                "provinces": province_columns,
                "rows": category_province_rows,
            }

        detail_out.append(block)

        # Among this Bookmark's over-SLA tickets, break down by PROVINCE how
        # many are still missing each of the four fields a lead needs to
        # answer CNO with (Group Problem, Action Team, รายละเอียด, Plan
        # Closed Date) - most affected province first, so gaps in triage
        # are visible at a glance.
        province_counts = {}
        for e in tickets:
            prov = str(e["PROVINCE"]).strip() or "(ไม่ระบุจังหวัด)"
            row = province_counts.setdefault(prov, {
                "missing_group_problem": 0, "missing_action_team": 0,
                "missing_detail": 0, "missing_plan_closed_date": 0,
            })
            if e["group_problem"] == UNSPECIFIED_GROUP_PROBLEM:
                row["missing_group_problem"] += 1
            if not str(e["action_team"]).strip():
                row["missing_action_team"] += 1
            if not str(e["detail"]).strip():
                row["missing_detail"] += 1
            if not str(e["plan_closed_date"]).strip():
                row["missing_plan_closed_date"] += 1
        province_rows = sorted(
            (
                {"province": p, **counts, "total_missing": sum(counts.values())}
                for p, counts in province_counts.items()
                if sum(counts.values()) > 0
            ),
            key=lambda x: (-x["missing_group_problem"], -x["total_missing"])
        )
        if province_rows:
            unspecified_by_province_out.append({"bookmark": bm, "provinces": province_rows})

        # Group Problem x Plan Closed Date matrix - who's stuck on what, and
        # when it's actually planned to close (or not planned at all yet).
        plan_dates = sorted({t["plan_closed_date"] for t in tickets if t["plan_closed_date"]})
        date_columns = plan_dates + [NO_PLAN_DATE]
        gp_plan_matrix = {}
        for e in tickets:
            col = e["plan_closed_date"] or NO_PLAN_DATE
            gp_plan_matrix.setdefault(e["group_problem"], {}).setdefault(col, 0)
            gp_plan_matrix[e["group_problem"]][col] += 1
        gp_plan_rows = []
        for gp, counts in gp_plan_matrix.items():
            row_total = sum(counts.values())
            gp_plan_rows.append({
                "group_problem": gp,
                "counts": {c: counts.get(c, 0) for c in date_columns},
                "total": row_total,
            })
        gp_plan_rows.sort(key=lambda r: -r["total"])
        if gp_plan_rows:
            group_problem_by_plan_date_out.append({
                "bookmark": bm, "date_columns": date_columns, "rows": gp_plan_rows,
            })

    return {
        "aging_order": AGING_ORDER,
        "over_aging_keys": OVER_24H_AGING_KEYS,
        "aging_colors": AGING_COLORS,
        "summary": summary_out,
        "province_summary": province_summary_out,
        "detail": detail_out,
        "unspecified_by_province": unspecified_by_province_out,
        "group_problem_by_plan_date": group_problem_by_plan_date_out,
        "online_remaining_hours_by_province": build_online_remaining_hours_by_province(entries),
        "total_over_sla": len(detail_entries),
        "generated_at": now_dt.strftime("%Y-%m-%d %H:%M:%S"),
        # insert_time is stamped by the EXTERNAL system (SCCD+ITSM) on the raw
        # sheet each time it syncs - different from generated_at above, which
        # is just when THIS response was built. Same field Pending Ticket
        # shows, so people can tell how fresh the underlying ticket data is.
        "insert_time": all_rows[0].get("insert_time") if all_rows else None,
        # Needed so Exclusive Pending / P0 Only can offer the same inline
        # edit dropdowns Pending Ticket uses, for the small set of people
        # allowed to edit from these pages too.
        "group_problem_options": GROUP_PROBLEM_OPTIONS,
        "action_team_options": ACTION_TEAM_OPTIONS,
    }


# ---------------------------------------------------------------------------
# P0 snapshot comparison: "P0 right now" vs "what P0 looked like at ~01:15
# today" (from the Drive backup file closest to that time), broken out by
# the same 4 groups as the Realtime Monitoring view buttons (Mobile SA1-4,
# Online/FBB SA1-4, NSA1-2, NSA3/4) - NOT the Exclusive Pending Bookmark
# grouping, which lumps NSA3/4 into "Others".
# ---------------------------------------------------------------------------

from datetime import time as _dtime, timedelta as _timedelta

_P0_SNAPSHOT_CACHE_TTL_SECONDS = 900  # 15 min - just for the (expensive) backup-file part, which is fixed once written and never changes
_p0_snapshot_cache = {"data": None, "ts": 0}
_p0_snapshot_lock = threading.Lock()

P0_COMPARISON_GROUPS = ["MB", "FBB", "NW_NSA12", "NSA34"]  # matches ticket_views.BOOKMARK_VIEWS keys, in display order

_p0_daily_trend_cache = {}
_p0_daily_trend_lock = threading.Lock()


def build_p0_daily_trend(gs_client, drive_service, days=7):
    """Returns {"dates": [...], "series": {group_key: [...counts...]}} - the
    same 'P0 count as it stood at ~01:15' snapshot classification the
    comparison cards already use, extended across the last `days` days
    instead of just yesterday, so each card can show a trend underneath.
    Each PAST day's count is cached forever once computed (a written
    backup file for a completed day never changes). TODAY is never
    cached at all - re-fetched and re-computed on every call. This is
    deliberate: an earlier version cached "today" as long as some file
    was found, but a stale result computed once (whatever its value -
    None OR a wrong/zero count from a since-fixed bug) would then be
    served for the rest of the day with no way to self-correct. Today's
    backup file also gets rewritten hourly by the external job, so unlike
    a completed past day there's no "final" version to lock in until the
    day is over anyway - re-fetching every time is both safer and correct.
    """
    from pending_trend import find_nightly_file, download_xlsx_as_rows

    today = bangkok_now().date()
    all_days = [today - _timedelta(days=i) for i in range(days - 1, -1, -1)]  # oldest -> newest

    dates = []
    series = {k: [] for k in P0_COMPARISON_GROUPS}
    debug_today = None
    for day in all_days:
        date_str = day.strftime("%Y-%m-%d")
        dates.append(date_str)
        is_today = day == today

        cached = None
        if not is_today:
            with _p0_daily_trend_lock:
                cached = _p0_daily_trend_cache.get(date_str)
        if cached is not None:
            counts = cached
        else:
            file_info = find_nightly_file(drive_service, day)
            if file_info is None:
                counts = {k: None for k in P0_COMPARISON_GROUPS}
            else:
                file_id, matched_dt, filename = file_info
                rows = download_xlsx_as_rows(drive_service, file_id)
                reference_dt = datetime.combine(day, _dtime(1, 15))
                counts = _count_p0_by_group(rows, reference_dt)
                if is_today:
                    debug_today = {"filename": filename, "matched_at": matched_dt.strftime("%Y-%m-%d %H:%M:%S"), "row_count": len(rows), "counts": counts}
            if not is_today:
                with _p0_daily_trend_lock:
                    _p0_daily_trend_cache[date_str] = counts

        for k in P0_COMPARISON_GROUPS:
            series[k].append(counts.get(k))

    return {"dates": dates, "series": series, "debug_today": debug_today}


def _classify_priority_at(target_finish_str, reference_dt):
    """Same P0/P1/P2 formula as realtime_monitor._classify_priority, but
    takes the reference point (normally "tomorrow's 01:15") directly rather
    than deriving it from "now" - needed here because we're classifying
    tickets as they stood at a FIXED past moment (a Drive snapshot), not
    live data."""
    tf = _parse_dt(target_finish_str)
    if tf is None:
        return None
    diff_hours = (reference_dt - tf).total_seconds() / 3600
    if diff_hours > 24:
        return "P0"
    elif diff_hours > 0:
        return "P1"
    else:
        return "P2"


def _count_p0_by_group(rows, reference_dt):
    from ticket_views import BOOKMARK_VIEWS, row_matches_view
    counts = {k: 0 for k in P0_COMPARISON_GROUPS}
    for r in rows:
        if str(r.get("Region", "")).strip() not in PENDING_TICKET_REGIONS:
            continue
        if str(r.get("SEVERITY", "")).strip() not in ALLOWED_SEVERITIES:
            continue
        priority = _classify_priority_at(r.get("TARGETFINISH"), reference_dt)
        for key in P0_COMPARISON_GROUPS:
            if not row_matches_view(r, key):
                continue
            # Same exception as the P0 Only detail/matrix view: the FBB
            # (Online, "4.FBB with SA1-4") group counts P0+P1 together;
            # every other group stays strict P0-only, unchanged.
            allowed = {"P0", "P1"} if key == "FBB" else {"P0"}
            if priority in allowed:
                counts[key] += 1
    return counts


def build_p0_snapshot_comparison(gs_client, drive_service, use_cache=True):
    """Returns {"snapshot_date", "snapshot_matched_at", "groups": [{"key",
    "label", "snapshot_p0", "current_p0", "diff"}, ...]} - P0 count right
    now vs P0 count as it stood at the ~01:15 Drive backup closest to
    today, for each of the 4 severity/bookmark groups. Falls back to
    yesterday's snapshot if today's isn't found yet (e.g. very early in the
    morning before the day's 01:15 backup has run).

    Only the snapshot half (the Drive backup download+count) is cached -
    that data is fixed once written and genuinely expensive to redo. The
    "current" half is recomputed on every call: fetch_live_rows already has
    its own short-TTL cache, so this stays cheap while never showing a
    current_p0 that's stale relative to what the ticket detail table shows
    right below it on the same page - the two used to share one 15-minute
    cache, which could make this card's count visibly disagree with the
    live detail list for up to 15 minutes after something changed."""
    from pending_trend import find_nightly_file, download_xlsx_as_rows
    from ticket_views import BOOKMARK_VIEWS

    now = time.monotonic()
    cached_snapshot = None
    if use_cache:
        with _p0_snapshot_lock:
            if _p0_snapshot_cache["data"] is not None and (now - _p0_snapshot_cache["ts"]) < _P0_SNAPSHOT_CACHE_TTL_SECONDS:
                cached_snapshot = _p0_snapshot_cache["data"]

    if cached_snapshot is not None:
        snapshot_date, matched_dt, filename, snapshot_counts, snapshot_row_count = cached_snapshot
    else:
        today = bangkok_now().date()
        file_info = find_nightly_file(drive_service, today)
        snapshot_date = today
        if file_info is None:
            file_info = find_nightly_file(drive_service, today - _timedelta(days=1))
            snapshot_date = today - _timedelta(days=1)
        if file_info is None:
            raise ValueError(f"No backup snapshot found near 01:15 for {today} or {today - _timedelta(days=1)}")

        file_id, matched_dt, filename = file_info
        snapshot_rows = download_xlsx_as_rows(drive_service, file_id)
        snapshot_row_count = len(snapshot_rows)
        reference_dt = datetime.combine(snapshot_date, _dtime(1, 15))
        snapshot_counts = _count_p0_by_group(snapshot_rows, reference_dt)
        if use_cache:
            with _p0_snapshot_lock:
                _p0_snapshot_cache["data"] = (snapshot_date, matched_dt, filename, snapshot_counts, snapshot_row_count)
                _p0_snapshot_cache["ts"] = now

    live_rows = fetch_live_rows(gs_client)
    now_dt = bangkok_now()
    current_reference_dt = (now_dt + _timedelta(days=1)).replace(hour=1, minute=15, second=0, microsecond=0)
    current_counts = _count_p0_by_group(live_rows, current_reference_dt)

    groups = []
    for key in P0_COMPARISON_GROUPS:
        s = snapshot_counts[key]
        c = current_counts[key]
        groups.append({
            "key": key, "label": BOOKMARK_VIEWS[key]["label"],
            "snapshot_p0": s, "current_p0": c, "diff": c - s,
        })

    return {
        "snapshot_date": snapshot_date.strftime("%Y-%m-%d"),
        "snapshot_matched_at": matched_dt.strftime("%Y-%m-%d %H:%M:%S"),
        "snapshot_filename": filename,
        "snapshot_row_count": snapshot_row_count,  # diagnostic - if this is 0, the backup file itself had no rows; if >0 but every snapshot_p0 is 0, the issue is in classification/matching, not the file
        "current_generated_at": now_dt.strftime("%Y-%m-%d %H:%M:%S"),
        "groups": groups,
    }


# ── Online SA1-4 SLA Monitoring ─────────────────────────────────────────
# New tab built on the exact same data source, Region/Severity filtering,
# and P0/P1 priority formula as Focus Priority 0 above (_classify_priority,
# imported from realtime_monitor - the "tomorrow 01:15" reference point is
# NOT reimplemented here, just reused) - scoped to the "Online" bookmark
# ("4.FBB with SA1-4") and P0+P1 together. Adds SLA Progress % (elapsed
# time / total CREATIONDATE-to-TARGETFINISH duration) and a hazard-hours
# Early Warning window - neither exists anywhere else in this codebase,
# since the existing Aging_Flag_Group is a different concept (buckets by
# how overdue an ALREADY-late ticket is, not by progress toward a
# not-yet-due deadline). This system has no CLOSEDTIME field anywhere and
# fetch_live_rows() only ever returns currently-pending tickets, so every
# entry here is implicitly "still open" by construction - confirmed
# acceptable, no closed-ticket tracking needed.

ONLINE_BOOKMARK_RAW = "4.FBB with SA1-4"


def _format_sla_duration_label(hours):
    """Displays SLA duration at whatever precision is actually meaningful
    - a 4-hour or 12-hour SLA needs to show as "4 Hours"/"12 Hours", not
    get rounded away to "0 Days" (which is what happened when this
    grouped by whole days only). Clean day-aligned multiples (24h, 72h,
    ...) show as "N Days" for readability; anything else shows in hours."""
    if hours is None:
        return None
    if hours >= 24 and hours % 24 == 0:
        return f"{int(hours // 24)} Days"
    return f"{int(hours)} Hours"


def _sla_risk_bucket(now_dt, target_dt, progress_pct):
    """🔴 over / 🟠 >=75% / 🟡 >=50% / 🔵 normal (<50%) - Over is decided
    by the clock (now past TARGETFINISH), matching how priority P0/P1
    already treat lateness; the other 3 buckets split the SLA Progress %
    range. Returns None if progress_pct couldn't be computed (missing/
    invalid CREATIONDATE or TARGETFINISH) - never guesses a bucket for
    that ticket rather than misclassifying it."""
    if target_dt and now_dt > target_dt:
        return "over"
    if progress_pct is None:
        return None
    if progress_pct >= 75:
        return "risk75"
    if progress_pct >= 50:
        return "risk50"
    return "normal"


def build_online_sla_response(gs_client=None):
    """Online SA1-4 SLA Monitoring - see module-level comment above for
    what's reused vs new. Returns entries (ticket-level, sorted by
    remaining SLA time ascending - most urgent first) for EVERY priority
    (P0/P1/P2 all included, not just P0+P1), KPI counts (including a
    priority breakdown so "everything together" stays readable), a
    Region->Province risk table, and an SLA-Duration distribution (+ x
    Province) table. SLA Duration groups are DISCOVERED from real data at
    HOUR precision (never hardcoded, never rounded away to whole days -
    a 4h or 12h SLA needs to stay visible as such)."""
    if gs_client is None:
        _, gs_client = get_drive_and_sheets_clients()

    all_rows = fetch_live_rows(gs_client)
    now_dt = bangkok_now()
    scoped = [
        r for r in all_rows
        if str(r.get("Region", "")).strip() in PENDING_TICKET_REGIONS
        and str(r.get("SEVERITY", "")).strip() in ALLOWED_SEVERITIES
        and str(r.get("Bookmark", "")).strip() == ONLINE_BOOKMARK_RAW
    ]
    work_log = load_work_log(gs_client)
    try:
        mateline_lookup = build_mateline_status_lookup(gs_client, now_dt.strftime("%Y-%m-%d"))
    except Exception:
        log.exception("GGS Daily mateline lookup failed for Online SLA - falling back to empty for this response")
        mateline_lookup = {}

    entries = []
    for r in scoped:
        priority = _classify_priority(r.get("TARGETFINISH"), now_dt)
        if priority is None:
            continue  # unparseable TARGETFINISH - can't classify at all, not even as P2

        creation_dt = _parse_dt(r.get("CREATIONDATE"))
        target_dt = _parse_dt(r.get("TARGETFINISH"))

        sla_duration_hours = None
        sla_duration_label = None
        sla_progress_pct = None
        remaining_hours = None
        if creation_dt and target_dt:
            total_seconds = (target_dt - creation_dt).total_seconds()
            if total_seconds > 0:
                sla_duration_hours = round(total_seconds / 3600)
                sla_duration_label = _format_sla_duration_label(sla_duration_hours)
                elapsed_seconds = (now_dt - creation_dt).total_seconds()
                sla_progress_pct = round(max(0, elapsed_seconds) / total_seconds * 100, 1)
            remaining_hours = round((target_dt - now_dt).total_seconds() / 3600, 2)

        sla_risk = _sla_risk_bucket(now_dt, target_dt, sla_progress_pct)

        ticket_id = str(r.get("TICKETID", "")).strip()
        wl = work_log.get(ticket_id, {})
        mateline = mateline_lookup.get(ticket_id.upper()) or {
            "status_mateline": "(ไม่พบใน MatelineX)", "mateline_wo_status": "",
        }
        over_sla_day = r.get("Over_SLA_Day")
        try:
            over_sla_day = float(over_sla_day)
        except (TypeError, ValueError):
            over_sla_day = 0

        entries.append({
            "TICKETID": ticket_id,
            "SUBJECT": r.get("SUBJECT", ""),
            "CINAME": r.get("CINAME", ""),
            "DISTRICT": r.get("DISTRICT", ""),
            "PROVINCE": str(r.get("PROVINCE", "")).strip() or "(ไม่ระบุ)",
            "Region": r.get("Region", ""),
            "TRUEOWNERGROUP": r.get("TRUEOWNERGROUP", ""),
            "priority": priority,
            "Aging_Flag_Group": str(r.get("Aging_Flag_Group", "")).strip() or UNSPECIFIED_AGING,
            "CREATIONDATE": r.get("CREATIONDATE", ""),
            "TARGETFINISH": r.get("TARGETFINISH", ""),
            "sla_duration_hours": sla_duration_hours,
            "sla_duration_label": sla_duration_label,
            "sla_progress_pct": sla_progress_pct,
            "sla_risk": sla_risk,  # 'over' | 'risk75' | 'risk50' | 'normal' | None
            "remaining_hours": remaining_hours,
            "early_warning_2h": remaining_hours is not None and 0 <= remaining_hours <= 2,
            "early_warning_6h": remaining_hours is not None and 0 <= remaining_hours <= 6,
            "early_warning_12h": remaining_hours is not None and 0 <= remaining_hours <= 12,
            "over_sla_day": over_sla_day,
            "status_mateline": mateline["status_mateline"],
            "mateline_wo_status": mateline["mateline_wo_status"],
            "group_problem": wl.get("group_problem") or UNSPECIFIED_GROUP_PROBLEM,
            "action_team": wl.get("action_team", ""),
        })

    # Most urgent (soonest TARGETFINISH / least remaining time) first -
    # entries with no computable remaining_hours sort last, not first.
    entries.sort(key=lambda e: e["remaining_hours"] if e["remaining_hours"] is not None else float("inf"))

    total = len(entries)
    risk_counts = {"over": 0, "risk75": 0, "risk50": 0, "normal": 0}
    priority_counts = {"P0": 0, "P1": 0, "P2": 0}
    for e in entries:
        if e["sla_risk"] in risk_counts:
            risk_counts[e["sla_risk"]] += 1
        if e["priority"] in priority_counts:
            priority_counts[e["priority"]] += 1
    ew_counts = {
        "2h": sum(1 for e in entries if e["early_warning_2h"]),
        "6h": sum(1 for e in entries if e["early_warning_6h"]),
        "12h": sum(1 for e in entries if e["early_warning_12h"]),
    }

    # Region -> Province risk table.
    province_stats = {}
    for e in entries:
        p = province_stats.setdefault(e["PROVINCE"], {
            "province": e["PROVINCE"], "region": e["Region"],
            "total": 0, "over": 0, "risk75": 0, "risk50": 0, "normal": 0,
        })
        p["total"] += 1
        if e["sla_risk"] in p:
            p[e["sla_risk"]] += 1
    province_table = sorted(province_stats.values(), key=lambda r: -r["total"])
    critical_province = max(province_table, key=lambda r: r["over"])["province"] if any(r["over"] for r in province_table) else None

    # SLA Duration distribution - groups DISCOVERED from real data at HOUR
    # precision, not a hardcoded list, and not rounded away to whole days.
    duration_counts = {}
    for e in entries:
        key = e["sla_duration_label"] or "(ไม่ระบุ)"
        duration_counts[key] = duration_counts.get(key, 0) + 1
    duration_table = sorted(
        [{"sla_duration": k, "ticket_count": v} for k, v in duration_counts.items()],
        key=lambda r: -r["ticket_count"],
    )

    # SLA Duration x Province.
    dp_stats = {}
    for e in entries:
        dkey = e["sla_duration_label"] or "(ไม่ระบุ)"
        key = (dkey, e["PROVINCE"])
        dp = dp_stats.setdefault(key, {
            "sla_duration": dkey, "province": e["PROVINCE"],
            "total": 0, "over": 0, "risk75": 0, "risk50": 0,
        })
        dp["total"] += 1
        if e["sla_risk"] in dp:
            dp[e["sla_risk"]] += 1
    duration_province_table = sorted(dp_stats.values(), key=lambda r: -r["total"])

    return {
        "entries": entries,
        "kpi": {
            "total": total,
            "over_sla": risk_counts["over"], "risk_75": risk_counts["risk75"],
            "risk_50": risk_counts["risk50"], "normal": risk_counts["normal"],
            "early_warning_2h": ew_counts["2h"], "early_warning_6h": ew_counts["6h"], "early_warning_12h": ew_counts["12h"],
            "critical_province": critical_province,
            "priority_p0": priority_counts["P0"], "priority_p1": priority_counts["P1"], "priority_p2": priority_counts["P2"],
        },
        "province_table": province_table,
        "duration_table": duration_table,
        "duration_province_table": duration_province_table,
        "insert_time": scoped[0].get("insert_time") if scoped else None,
    }
